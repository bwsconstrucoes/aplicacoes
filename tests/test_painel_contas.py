# -*- coding: utf-8 -*-
"""
As contas do painel que não são SQL: rateio, divisão entre sócios, simulação
de caixa e exportação.

Estas são as regras que mais custam caro quando erram, porque erram calado: um
rateio torto não quebra tela nenhuma, só distribui o lucro errado entre sócios.
Por isso os testes aqui verificam **invariantes** — coisas que têm de ser
verdade sempre — e não só "a função rodou":

  - rateio move custo, não cria nem apaga: o que sai do administrativo tem de
    chegar nas obras (ou aparecer como sobra, com o motivo);
  - a soma das quotas dos sócios fecha com o resultado do projeto, inclusive
    quando há sócio externo, que é o caso onde a conta muda;
  - o caixa acumulado é a soma corrida dos meses, sem buraco em mês vazio.

Nenhum teste aqui abre banco: as funções recebem listas e devolvem listas.
"""
from __future__ import annotations

import datetime as dt
import io
import json

import pytest

from app.apps.painel import excel, prestacao, simulacao


# ===========================================================================
# 1. Rateio administrativo
# ===========================================================================
CONFIG = {
    "projeto_matriz": "MATRIZ_PROJ",
    "depto_admin_matriz": "ADM MATRIZ",
    "depto_admin_filial": "ADM FILIAL",
    "grupo_pessoal": "Despesas com Pessoal",
    "taxa_adm_pct": "1.5",
    "residual": "1",
}

# duas obras da matriz e uma da filial
APURACAO = [
    {"mes": "2025-01", "obra": "CASA", "projeto": "MATRIZ_PROJ",
     "receita_liquida": 10000.0, "retencoes": 500.0, "despesas": -4000.0},
    {"mes": "2025-01", "obra": "PREDIO", "projeto": "MATRIZ_PROJ",
     "receita_liquida": 6000.0, "retencoes": 0.0, "despesas": -2000.0},
    {"mes": "2025-01", "obra": "PONTE", "projeto": "NORDESTE",
     "receita_liquida": 8000.0, "retencoes": 0.0, "despesas": -3000.0},
    {"mes": "2025-01", "obra": "ADM MATRIZ", "projeto": "MATRIZ_PROJ",
     "receita_liquida": 0.0, "retencoes": 0.0, "despesas": -1000.0},
]

# pessoal: CASA 300, PREDIO 100, PONTE 400  -> pesos 3:1 na matriz
PESSOAL = [("2025-01", "CASA", 300.0), ("2025-01", "PREDIO", 100.0),
           ("2025-01", "PONTE", 400.0)]

ADMIN = [
    {"mes": "2025-01", "depto": "ADM MATRIZ", "grupo": "Despesas com Pessoal",
     "categoria": "Salários", "valor": -800.0},
    {"mes": "2025-01", "depto": "ADM MATRIZ", "grupo": "Despesas Administrativas",
     "categoria": "Aluguéis", "valor": -200.0},
]


def _obras():
    return prestacao.classificar_obras(APURACAO, CONFIG)


def test_administrativo_nao_e_obra():
    """O departamento administrativo é a estrutura a ser rateada. Se ele
    entrasse como obra, receberia rateio de si mesmo."""
    obras = _obras()
    assert set(obras) == {"CASA", "PREDIO", "PONTE"}
    assert obras["CASA"]["lado"] == prestacao.MATRIZ
    assert obras["PONTE"]["lado"] == prestacao.FILIAL


def test_rateio_segue_a_proporcao_do_pessoal():
    """Metade dos 800 de pessoal da matriz vai para as obras da matriz, na
    proporção 300:100 — logo 300 para a CASA e 100 para o PRÉDIO."""
    regra = {"nome": "metade do pessoal", "depto": "ADM MATRIZ", "todas": 0,
             "grupos": json.dumps(["Despesas com Pessoal"]), "categorias": "[]",
             "pct": 50, "escopo": prestacao.MATRIZ, "mes_ini": "", "mes_fim": "",
             "ativo": 1}
    resultado = prestacao.calcular_rateio(ADMIN, PESSOAL, _obras(), [regra],
                                          {**CONFIG, "residual": "0"})
    aloc = resultado["alocacoes"]
    assert aloc[("CASA", "2025-01")] == pytest.approx(-300.0)
    assert aloc[("PREDIO", "2025-01")] == pytest.approx(-100.0)
    assert ("PONTE", "2025-01") not in aloc      # é da filial, fora do escopo


def test_o_residuo_impede_que_custo_desapareca():
    """Esta é a invariante que mais importa: rateio MOVE custo, não o apaga.

    Com o resíduo ligado, o total distribuído tem de bater com o total do
    administrativo. Sem isso, o resultado da empresa mudaria só por causa de
    como o rateio foi configurado — o que seria uma mentira."""
    regra = {"nome": "só o aluguel", "depto": "ADM MATRIZ", "todas": 0,
             "grupos": "[]", "categorias": json.dumps(["Aluguéis"]),
             "pct": 100, "escopo": "AMBAS", "mes_ini": "", "mes_fim": "",
             "ativo": 1}
    resultado = prestacao.calcular_rateio(ADMIN, PESSOAL, _obras(), [regra], CONFIG)
    distribuido = sum(resultado["alocacoes"].values())
    sobrou = sum(s["valor"] for s in resultado["sobras"])
    total_admin = sum(linha["valor"] for linha in ADMIN)
    assert distribuido + sobrou == pytest.approx(total_admin, abs=0.01)


def test_vigencia_da_regra_e_respeitada():
    """'Isto foi compartilhado até outubro de 2025' tem de parar em outubro."""
    admin = ADMIN + [{"mes": "2025-11", "depto": "ADM MATRIZ",
                      "grupo": "Despesas com Pessoal", "categoria": "Salários",
                      "valor": -500.0}]
    pessoal = PESSOAL + [("2025-11", "CASA", 100.0)]
    regra = {"nome": "até outubro", "depto": "ADM MATRIZ", "todas": 1,
             "grupos": "[]", "categorias": "[]", "pct": 100, "escopo": "AMBAS",
             "mes_ini": "", "mes_fim": "2025-10", "ativo": 1}
    resultado = prestacao.calcular_rateio(admin, pessoal, _obras(), [regra],
                                          {**CONFIG, "residual": "0"})
    meses = {mes for _obra, mes in resultado["alocacoes"]}
    assert meses == {"2025-01"}


def test_regra_inativa_nao_pega_nada():
    regra = {"nome": "desligada", "depto": "ADM MATRIZ", "todas": 1,
             "grupos": "[]", "categorias": "[]", "pct": 100, "escopo": "AMBAS",
             "mes_ini": "", "mes_fim": "", "ativo": 0}
    resultado = prestacao.calcular_rateio(ADMIN, PESSOAL, _obras(), [regra],
                                          {**CONFIG, "residual": "0"})
    assert resultado["alocacoes"] == {}


def test_sem_pessoal_no_mes_o_custo_vira_sobra_com_motivo():
    """Não dá para ratear proporcionalmente a nada. O custo então aparece como
    sobra — visível, com o motivo — em vez de sumir da conta."""
    resultado = prestacao.calcular_rateio(
        ADMIN, [], _obras(),
        [{"nome": "tudo", "depto": "ADM MATRIZ", "todas": 1, "grupos": "[]",
          "categorias": "[]", "pct": 100, "escopo": "AMBAS", "mes_ini": "",
          "mes_fim": "", "ativo": 1}],
        {**CONFIG, "residual": "0"})
    assert resultado["alocacoes"] == {}
    assert sum(s["valor"] for s in resultado["sobras"]) == pytest.approx(-1000.0)
    assert all(s["motivo"] for s in resultado["sobras"])


# ===========================================================================
# 2. A divisão entre sócios
# ===========================================================================
def test_projeto_so_da_bws_divide_o_resultado_com_rateio():
    por_projeto = {"OBRA1": {"receita_bruta": 10000.0, "receita_liquida": 9500.0,
                             "retencoes": 500.0, "despesas": -4000.0,
                             "rateio": -500.0, "resultado_direto": 5500.0,
                             "resultado": 5000.0}}
    quotas = prestacao.quotas_por_socio(
        por_projeto,
        [{"projeto": "OBRA1", "socio": "ANA", "tipo": "Interno", "pct": 60},
         {"projeto": "OBRA1", "socio": "BENTO", "tipo": "Interno", "pct": 40}],
        CONFIG)
    assert {q["socio"]: q["quota"] for q in quotas} == {"ANA": 3000.0, "BENTO": 2000.0}
    assert sum(q["quota"] for q in quotas) == pytest.approx(5000.0)


def test_com_socio_externo_a_soma_das_quotas_ainda_fecha():
    """A invariante que sustenta a tela inteira.

    Com parceiro externo a conta muda: cobra-se uma taxa de administração da
    parceria, todos dividem o resultado direto menos a taxa, e a taxa mais o
    rateio da obra voltam só para os sócios da BWS. Se essa volta não for
    exata, dinheiro aparece ou some — e ninguém percebe olhando a tela."""
    numeros = {"receita_bruta": 100000.0, "receita_liquida": 95000.0,
               "retencoes": 5000.0, "despesas": -60000.0,
               "rateio": -4000.0, "resultado_direto": 35000.0,
               "resultado": 31000.0}
    quotas = prestacao.quotas_por_socio(
        {"PARCERIA": numeros},
        [{"projeto": "PARCERIA", "socio": "BWS-ANA", "tipo": "Interno", "pct": 30},
         {"projeto": "PARCERIA", "socio": "BWS-BENTO", "tipo": "Interno", "pct": 20},
         {"projeto": "PARCERIA", "socio": "PARCEIRO", "tipo": "Externo", "pct": 50}],
        CONFIG)
    assert sum(q["quota"] for q in quotas) == pytest.approx(numeros["resultado"], abs=0.01)

    # o externo divide a base da parceria, sem taxa nem rateio
    externo = next(q for q in quotas if q["tipo"] == "Externo")
    taxa = 0.015 * numeros["receita_bruta"]
    assert externo["base"] == pytest.approx(numeros["resultado_direto"] - taxa)
    assert externo["credito_bws"] == 0.0

    # os internos recebem de volta a taxa e o rateio, na proporção entre eles
    internos = [q for q in quotas if q["tipo"] == "Interno"]
    assert sum(q["credito_bws"] for q in internos) == pytest.approx(
        taxa + numeros["rateio"], abs=0.01)


def test_o_externo_nao_paga_a_estrutura_da_construtora():
    """É a razão de a conta ser diferente: ele entrou na obra, não na BWS."""
    numeros = {"receita_bruta": 0.0, "receita_liquida": 0.0, "retencoes": 0.0,
               "despesas": 0.0, "rateio": -10000.0, "resultado_direto": 0.0,
               "resultado": -10000.0}
    quotas = prestacao.quotas_por_socio(
        {"PARCERIA": numeros},
        [{"projeto": "PARCERIA", "socio": "BWS", "tipo": "Interno", "pct": 50},
         {"projeto": "PARCERIA", "socio": "PARCEIRO", "tipo": "Externo", "pct": 50}],
        CONFIG)
    externo = next(q for q in quotas if q["tipo"] == "Externo")
    interno = next(q for q in quotas if q["tipo"] == "Interno")
    assert externo["quota"] == pytest.approx(0.0)          # nada do rateio
    assert interno["quota"] == pytest.approx(-10000.0)     # o rateio inteiro


def test_ajustes_tem_o_sinal_certo():
    """Trocar o sinal aqui inverte quem deve a quem."""
    assert prestacao.efeito_do_ajuste("Valor Percebido (-)", 1000) == -1000
    assert prestacao.efeito_do_ajuste("Valor Percebido (-)", -1000) == -1000
    assert prestacao.efeito_do_ajuste("Dívida Assumida (+)", 500) == 500
    assert prestacao.efeito_do_ajuste("Dívida Assumida (+)", -500) == 500
    assert prestacao.efeito_do_ajuste("Outro (+/-)", -300) == -300


def test_posicao_desconta_o_que_o_socio_ja_tirou():
    quotas = [{"socio": "ANA", "tipo": "Interno", "quota": 5000.0}]
    ajustes = [{"socio": "ANA", "tipo": "Valor Percebido (-)", "valor": 2000.0}]
    posicao = prestacao.posicao_dos_socios(quotas, ajustes)
    assert posicao[0]["saldo"] == pytest.approx(3000.0)


# ===========================================================================
# 3. A simulação de caixa
# ===========================================================================
def test_mes_vazio_no_meio_nao_vira_buraco():
    """Se um mês sem movimento sumisse, o acumulado daria um salto no gráfico
    e pareceria que o caixa mudou de patamar do nada."""
    meses = simulacao.meses_do_periodo(dt.date(2024, 11, 1), dt.date(2025, 2, 1))
    assert [m.strftime("%m/%Y") for m in meses] == \
        ["11/2024", "12/2024", "01/2025", "02/2025"]


def test_projeto_escolhido_puxa_todas_as_suas_obras():
    mapa = {"CASA": "ALFA", "PREDIO": "ALFA", "PONTE": "BETA"}
    pesos = simulacao.pesos_do_conjunto([("projeto:ALFA", 100)], mapa,
                                        ["CASA", "PREDIO", "PONTE"])
    assert pesos == {"CASA": 1.0, "PREDIO": 1.0, "PONTE": 0.0}


def test_a_mesma_obra_em_duas_linhas_soma_mas_para_em_cem():
    mapa = {"CASA": "ALFA"}
    pesos = simulacao.pesos_do_conjunto(
        [("obra:CASA", 60), ("obra:CASA", 70)], mapa, ["CASA"])
    assert pesos["CASA"] == 1.0


def test_simulacao_separa_o_conjunto_do_resto():
    linhas = [(dt.date(2025, 1, 1), "CASA", -1000.0),
              (dt.date(2025, 1, 1), "PONTE", 3000.0),
              (dt.date(2025, 2, 1), "CASA", -500.0)]
    financeiro = [{"mes": dt.date(2025, 1, 1), "emprestimo_tomado": 2000.0,
                   "emprestimo_pago": 0.0, "aporte_recebido": 0.0,
                   "dividendo_pago": 0.0, "outros": 0.0}]
    r = simulacao.simular(linhas, financeiro, [("obra:CASA", 100)],
                          {"CASA": "ALFA", "PONTE": "BETA"})
    ultimo = r["linhas"][-1]
    assert ultimo["conjunto_a"] == pytest.approx(-1500.0)
    assert ultimo["resto"] == pytest.approx(3000.0)
    assert ultimo["empresa"] == pytest.approx(1500.0)
    # o caixa reconstruído inclui o empréstimo, a linha da empresa não
    assert ultimo["caixa_reconstruido"] == pytest.approx(3500.0)


def test_a_leitura_avisa_quando_a_empresa_precisou_do_banco():
    linhas = [(dt.date(2025, 1, 1), "CASA", -5000.0)]
    financeiro = [{"mes": dt.date(2025, 1, 1), "emprestimo_tomado": 5000.0,
                   "emprestimo_pago": 0.0, "aporte_recebido": 0.0,
                   "dividendo_pago": 0.0, "outros": 0.0}]
    r = simulacao.simular(linhas, financeiro, [("obra:CASA", 100)], {"CASA": "ALFA"})
    texto = " ".join(r["leitura"])
    assert "precisou do banco" in texto
    assert "não precisou" not in texto


def test_a_leitura_denuncia_caixa_negativo():
    """Caixa não fica negativo na vida real. Se ficar, falta uma fonte de
    dinheiro na base — e a tela tem de dizer isso, não maquiar."""
    linhas = [(dt.date(2025, 1, 1), "CASA", -5000.0)]
    r = simulacao.simular(linhas, [], [("obra:CASA", 100)], {"CASA": "ALFA"})
    texto = " ".join(r["leitura"])
    assert "não fechou" in texto
    assert "investigar" in texto


def test_saldo_inicial_levanta_a_linha_do_caixa():
    linhas = [(dt.date(2025, 1, 1), "CASA", -5000.0)]
    r = simulacao.simular(linhas, [], [("obra:CASA", 100)], {"CASA": "ALFA"},
                          saldo_inicial=8000.0)
    assert r["linhas"][-1]["caixa_reconstruido"] == pytest.approx(3000.0)
    assert "passou" in " ".join(r["leitura"])


# ===========================================================================
# 4. A exportação
# ===========================================================================
def test_a_planilha_marca_dinheiro_como_dinheiro():
    """Sem o formato de moeda a coluna vira número solto e ninguém soma nada
    olhando. Quem decide é o TÍTULO da coluna, não o valor."""
    from openpyxl import load_workbook

    livro = load_workbook(io.BytesIO(excel.montar(
        [("Teste", [("nome", "Descrição"), ("valor", "Valor")],
          [{"nome": "Manutenção", "valor": -1234.5}])])))
    folha = livro["Teste"]
    assert folha["A1"].value == "Descrição"
    assert folha["B2"].value == -1234.5              # número, não texto
    assert "R$" in folha["B2"].number_format


def test_data_na_planilha_sai_no_formato_brasileiro():
    from openpyxl import load_workbook

    livro = load_workbook(io.BytesIO(excel.montar(
        [("Teste", [("data", "Data")], [{"data": dt.date(2025, 3, 9)}])])))
    celula = livro["Teste"]["A2"]
    assert celula.value == dt.datetime(2025, 3, 9, 0, 0)
    assert celula.number_format == "DD/MM/YYYY"


def test_campo_ausente_vira_celula_vazia_e_nao_erro():
    from openpyxl import load_workbook

    livro = load_workbook(io.BytesIO(excel.montar(
        [("Teste", [("a", "A"), ("b", "B")], [{"a": 1}])])))
    assert livro["Teste"]["B2"].value is None


def test_aba_vazia_entra_com_aviso_em_vez_de_sumir():
    """Aba que some deixa quem abre sem saber se não havia dado ou se o
    relatório falhou na metade."""
    from openpyxl import load_workbook

    livro = load_workbook(io.BytesIO(excel.montar(
        [("Top Credores", [("nome", "Credor")], [])])))
    assert livro.sheetnames == ["Top Credores"]
    assert "sem dados" in livro["Top Credores"]["A1"].value


def test_nome_de_aba_longo_ou_proibido_nao_derruba_o_arquivo():
    r"""O Excel recusa aba com mais de 31 caracteres ou com : \ / ? * [ ] —
    e um arquivo recusado é pior do que uma aba com nome cortado."""
    from openpyxl import load_workbook

    nome = "Comprometido x Executado / por obra [2025]"
    livro = load_workbook(io.BytesIO(excel.montar(
        [(nome, [("a", "A")], [{"a": 1}])])))
    (saida,) = livro.sheetnames
    assert len(saida) <= 31
    assert not set(saida) & set(r':\/?*[]')


# ===========================================================================
# 5. A hora que aparece na tela
# ===========================================================================
# O servidor roda em UTC. Sem converter, uma carga das 13h29 aparece como
# 16h29 — e quem lê acha que o relógio quebrou, ou que aconteceu outra coisa.

def test_hora_do_servidor_vira_hora_de_brasilia():
    from app.apps.painel import horario
    utc = dt.datetime(2026, 9, 2, 16, 29, tzinfo=dt.timezone.utc)
    assert horario.texto(utc) == "02/09/2026 às 13:29"


def test_data_sem_fuso_e_tratada_como_utc():
    """Tratar como hora local deslocaria de novo, na direção errada — e o erro
    passaria despercebido porque o número continua parecendo uma hora."""
    from app.apps.painel import horario
    assert horario.texto(dt.datetime(2026, 9, 2, 16, 29)) == "02/09/2026 às 13:29"


def test_sem_data_mostra_travessao_e_nao_None():
    from app.apps.painel import horario
    assert horario.texto(None) == "—"


def test_a_virada_do_dia_e_respeitada():
    """01:00 UTC ainda é o dia anterior no Brasil. Errar isso joga um
    lançamento para o dia seguinte no relatório."""
    from app.apps.painel import horario
    meia_noite = dt.datetime(2026, 9, 3, 1, 0, tzinfo=dt.timezone.utc)
    assert horario.texto(meia_noite) == "02/09/2026 às 22:00"


# ===========================================================================
# 6. Cenários de rateio — ajustar e comparar antes de gravar
# ===========================================================================
# Mudar uma regra de rateio muda quanto de custo cai em cada obra e, por
# consequência, quanto cabe a cada sócio. O cenário existe para ver isso ANTES
# de gravar. O que se prova aqui:
#   - a alteração fica em memória: as regras gravadas não são tocadas;
#   - a comparação só mostra obra que mudou, e Δ negativo é obra que piorou;
#   - o total continua fechando: rateio move custo, o cenário não cria nem apaga;
#   - valor absurdo digitado é preso no limite, não vira erro na cara de quem
#     está simulando.
REGRA_BASE = {
    "id": 7, "nome": "Administrativo matriz", "depto": "ADM MATRIZ",
    "todas": 1, "grupos": "[]", "categorias": "[]", "pct": 100.0,
    "escopo": "AMBAS", "mes_ini": "", "mes_fim": "", "ativo": 1,
}


def _cenario(mudanca):
    return prestacao.regras_do_cenario([dict(REGRA_BASE)], {"7": mudanca})


def test_o_cenario_nao_toca_nas_regras_gravadas():
    """A promessa da tela é "nada é gravado enquanto você não mandar". Se o
    cálculo alterasse a lista recebida, a tela de Regras passaria a mostrar o
    cenário como se fosse o oficial."""
    gravadas = [dict(REGRA_BASE)]
    prestacao.regras_do_cenario(gravadas, {"7": {"pct": "40", "ativo": 0}})
    assert gravadas[0]["pct"] == 100.0
    assert gravadas[0]["ativo"] == 1


def test_regra_que_ninguem_mexeu_passa_inteira():
    """Um cenário que altera uma regra não pode alterar as outras trinta."""
    outra = dict(REGRA_BASE, id=8, nome="Filial", pct=55.0)
    saida = prestacao.regras_do_cenario([dict(REGRA_BASE), outra],
                                        {"7": {"pct": "40"}})
    assert saida[0]["pct"] == 40.0
    assert saida[1]["pct"] == 55.0


def test_porcentagem_absurda_e_presa_no_limite():
    """Quem está simulando digita rápido. 500% vira 100, e -30 vira 0 — a tela
    antiga fazia igual. Recusar com erro atrapalharia mais do que ajuda."""
    assert _cenario({"pct": "500"})[0]["pct"] == 100.0
    assert _cenario({"pct": "-30"})[0]["pct"] == 0.0
    # vírgula decimal, que é como se digita em português
    assert _cenario({"pct": "12,5"})[0]["pct"] == 12.5
    # texto sem sentido não derruba a tela: fica como estava
    assert _cenario({"pct": "abc"})[0]["pct"] == 100.0


def test_escopo_desconhecido_cai_em_ambas():
    assert _cenario({"escopo": "matriz"})[0]["escopo"] == "MATRIZ"
    assert _cenario({"escopo": "inventado"})[0]["escopo"] == "AMBAS"


def test_desligar_a_regra_no_cenario_para_de_ratear():
    """Regra desligada não pega nada — e, sem resíduo, o custo administrativo
    deixa de ser distribuído."""
    obras = _obras()
    config_sem_residuo = dict(CONFIG, residual="0")
    ligada = prestacao.calcular_rateio(ADMIN, PESSOAL, obras,
                                       [dict(REGRA_BASE)], config_sem_residuo)
    desligada = prestacao.calcular_rateio(ADMIN, PESSOAL, obras,
                                          _cenario({"ativo": 0}), config_sem_residuo)
    assert sum(ligada["alocacoes"].values()) != 0
    assert sum(desligada["alocacoes"].values()) == 0


def test_a_comparacao_mostra_so_a_obra_que_mudou():
    """Numa lista de cem obras, mostrar as noventa que continuam iguais esconde
    as dez que interessam."""
    obras = _obras()
    oficial = prestacao.calcular_rateio(ADMIN, PESSOAL, obras,
                                        [dict(REGRA_BASE)], CONFIG)
    # só a matriz: PONTE (filial) deixa de receber desta regra
    cenario = prestacao.calcular_rateio(ADMIN, PESSOAL, obras,
                                        _cenario({"escopo": "MATRIZ"}), CONFIG)
    comparacao = prestacao.comparar_por_obra(
        prestacao.apurar(APURACAO, obras, oficial["alocacoes"]),
        prestacao.apurar(APURACAO, obras, cenario["alocacoes"]))

    mudaram = {l["obra"] for l in comparacao}
    assert "PONTE" in mudaram
    for linha in comparacao:
        assert abs(linha["delta_resultado"]) > prestacao.LIMITE_DE_RUIDO


def test_delta_negativo_e_a_obra_que_passa_a_receber_mais_custo():
    """É contraintuitivo o bastante para estar escrito na tela — e para ter
    teste: mais rateio significa resultado MENOR, e o Δ tem de sair negativo.
    Se o sinal invertesse, a tela mostraria em verde a obra que piorou."""
    obras = _obras()
    sem_rateio = prestacao.apurar(APURACAO, obras, {})
    com_rateio = prestacao.apurar(
        APURACAO, obras,
        prestacao.calcular_rateio(ADMIN, PESSOAL, obras,
                                  [dict(REGRA_BASE)], CONFIG)["alocacoes"])

    comparacao = prestacao.comparar_por_obra(sem_rateio, com_rateio)
    assert comparacao, "o rateio muda alguma coisa"
    for linha in comparacao:
        # o rateio é custo (negativo): quem recebe mais, piora
        assert linha["delta_rateio"] < 0
        assert linha["delta_resultado"] < 0
    # a pior obra vem primeiro
    assert comparacao[0]["delta_resultado"] == min(
        l["delta_resultado"] for l in comparacao)


def test_o_cenario_move_custo_mas_nao_cria_nem_apaga():
    """A invariante que vale para o rateio vale para o cenário: o que sai do
    administrativo chega nas obras ou aparece como sobra. Um cenário que
    "melhora" tudo só porque perdeu dinheiro no caminho seria um erro caro."""
    obras = _obras()
    total_admin = sum(l["valor"] for l in ADMIN)
    for mudanca in ({"pct": "40"}, {"escopo": "MATRIZ"}, {"ativo": 0}):
        rateio = prestacao.calcular_rateio(ADMIN, PESSOAL, obras,
                                           _cenario(mudanca), CONFIG)
        distribuido = sum(rateio["alocacoes"].values())
        sobrou = sum(s["valor"] for s in rateio["sobras"])
        assert abs((distribuido + sobrou) - total_admin) < 0.01, mudanca


def test_o_resumo_conta_o_rateado_e_o_que_sobrou_dos_dois_lados():
    """A sobra importa tanto quanto o rateio: um cenário que rateia menos
    porque empurrou valor para a sobra não melhorou nada, e a tela precisa
    deixar isso visível."""
    obras = _obras()
    oficial = prestacao.calcular_rateio(ADMIN, PESSOAL, obras,
                                        [dict(REGRA_BASE)], CONFIG)
    cenario = prestacao.calcular_rateio(ADMIN, PESSOAL, obras,
                                        _cenario({"pct": "50"}), CONFIG)
    resumo = prestacao.resumo_do_cenario(oficial, cenario)
    assert resumo["delta_rateado"] == pytest.approx(
        resumo["rateado_cenario"] - resumo["rateado_oficial"])
    assert resumo["delta_sobra"] == pytest.approx(
        resumo["sobra_cenario"] - resumo["sobra_oficial"])


def test_sem_mexer_em_nada_a_tela_nao_diz_que_mexeu():
    """O formulário devolve TODOS os campos a cada Recalcular, inclusive os que
    ninguém tocou — e o banco devolve `pct` como número enquanto o formulário
    devolve texto. Comparar os dicionários direto acusaria alteração sempre."""
    gravadas = [dict(REGRA_BASE)]
    igual = prestacao.regras_do_cenario(
        gravadas, {"7": {"pct": "100.0", "escopo": "AMBAS",
                         "mes_ini": "", "mes_fim": "", "ativo": "on"}})
    assert not prestacao.cenario_difere(gravadas, igual)

    diferente = prestacao.regras_do_cenario(gravadas, {"7": {"pct": "99"}})
    assert prestacao.cenario_difere(gravadas, diferente)
