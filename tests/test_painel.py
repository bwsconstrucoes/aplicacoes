# -*- coding: utf-8 -*-
"""
Testes do Painel Financeiro OMIE.

Estes testes existem por um motivo específico. A suíte do ERP dubla a sessão e,
por isso, **não cobre SQL** (ver `conftest.py`). Mas o painel é quase todo SQL:
a conversão do painel antigo trocou "abrir o arquivo e somar em Python" por
"pedir a soma ao Postgres". Então o que precisa de rede aqui é justamente o que
a suíte existente não pega:

  - a tradução dos marcadores `?` para o Postgres, que passa por dentro de
    aspas e de sinais de porcentagem;
  - a ORDEM dos parâmetros no WHERE — um erro aqui não quebra nada, só devolve
    o número errado, que é bem pior;
  - a explosão de um título nas linhas por obra, com a linha separada de
    imposto retido;
  - a geometria dos gráficos.

Nada aqui abre banco: a conexão é dublada com dados em memória, no mesmo
espírito da `SessaoFalsa` do ERP.
"""
from __future__ import annotations

import datetime as dt
import io

import pytest

from app.apps.painel.db import traduzir_placeholders
from app.apps.painel.consultas import Filtros
from app.apps.painel import graficos
from app.apps.painel.sync import fato


# ===========================================================================
# 1. Tradução dos marcadores para o Postgres
# ===========================================================================
def test_marcadores_viram_placeholder_do_postgres():
    assert traduzir_placeholders("SELECT a FROM t WHERE b=? AND c=?") == \
        "SELECT a FROM t WHERE b=%s AND c=%s"


def test_porcentagem_literal_e_protegida():
    """`%` cru faria o psycopg2 tentar interpolar e estourar. Tem de virar `%%`.

    Isto não é teoria: a consulta que separa o imposto retido usa
    `categoria ILIKE '%Retido%'`."""
    assert traduzir_placeholders("SELECT a FROM t WHERE c ILIKE '%Retido%'") == \
        "SELECT a FROM t WHERE c ILIKE '%%Retido%%'"


def test_interrogacao_dentro_de_aspas_nao_e_marcador():
    """Uma interrogação em texto é dado, não parâmetro."""
    assert traduzir_placeholders("SELECT ? WHERE t = 'e ai?'") == \
        "SELECT %s WHERE t = 'e ai?'"


def test_consultas_do_painel_traduzem_sem_sobrar_marcador():
    """Varre as consultas reais: nenhuma pode ficar com `?` depois da tradução."""
    from app.apps.painel import consultas
    f = Filtros(anos=[2025], projetos=["OBRA1"], departamentos=["DEP"])
    where, _ = f.where("analise = 'DRE' AND tipo = ?", [consultas.REC])
    traduzida = traduzir_placeholders(f"SELECT SUM({consultas.COMPROMETIDO}) FROM fato{where}")
    assert "?" not in traduzida
    assert traduzida.count("%s") == 4          # 3 filtros + o tipo
    assert "'%%Retido%%'" not in traduzida     # esta consulta não usa retido
    assert "~*" in traduzida                   # a regra de "foi pago" continua lá


# ===========================================================================
# 2. Ordem dos parâmetros — o erro silencioso
# ===========================================================================
def test_ordem_dos_parametros_segue_a_ordem_do_where():
    """O parâmetro de `extra` tem de ficar DEPOIS dos filtros da barra lateral.

    Se trocar a ordem, o Postgres compara ano com texto e o painel mostra número
    errado sem reclamar de nada. É exatamente o tipo de defeito que a sessão
    dublada do ERP não pega."""
    f = Filtros(anos=[2024, 2025], projetos=["ALFA"], excluir_trf=True)
    where, params = f.where("tipo = ?", ["2. Contas a Pagar"])

    assert where.index("ano = ANY") < where.index("projeto = ANY") < where.index("tipo =")
    assert params == [[2024, 2025], ["ALFA"], "2. Contas a Pagar"]


def test_filtro_de_ano_preserva_titulos_sem_data():
    """Título em aberto sem data é backlog de hoje. Se o filtro de ano o
    descartasse, 'a pagar' e 'a receber' sumiriam da tela."""
    where, _ = Filtros(anos=[2025]).where()
    assert "data IS NULL" in where


def test_sem_filtro_nenhum_nao_gera_where():
    where, params = Filtros(excluir_trf=False).where()
    assert where == ""
    assert params == []


def test_transferencias_saem_por_padrao():
    """Transferência entre contas da própria empresa apareceria como entrada E
    saída do mesmo valor, inflando os dois lados."""
    where, _ = Filtros().where()
    assert "analise <> 'TRF'" in where


# ===========================================================================
# 3. A explosão do título em linhas por obra
# ===========================================================================
class ConexaoFalsa:
    """Conexão dublada: devolve as listas que o teste montou, sem banco nenhum.

    Só precisa responder o que `gerar_linhas_fato` pergunta — títulos (em
    stream), movimentos e rateios do bloco, e os catálogos."""

    def __init__(self, titulos, movimentos=(), rateios=(), categorias=(),
                 clientes=(), obras=(), contas=()):
        self.titulos = list(titulos)
        self.movimentos = list(movimentos)
        self.rateios = list(rateios)
        self.categorias = list(categorias)
        self.clientes = list(clientes)
        self.obras = list(obras)
        self.contas = list(contas)

    def _resposta(self, sql):
        # a ordem importa: "FROM categoria_de_para" contém "FROM cat"
        if "FROM categoria_de_para" in sql:
            return []
        if "FROM cat" in sql:
            return self.categorias
        if "FROM clientes" in sql:
            return self.clientes
        if "FROM depto_projeto" in sql:
            return self.obras
        if "FROM contas_correntes" in sql:
            return self.contas
        if "FROM movimentos" in sql:
            return self.movimentos
        if "FROM rateio" in sql:
            return self.rateios
        if "FROM titulos" in sql:
            return self.titulos
        raise AssertionError(f"consulta inesperada: {sql[:80]}")

    def execute(self, sql, params=()):
        return CursorFalso(self._resposta(sql))

    def executar_em_stream(self, sql, params=(), por_vez=2000):
        return CursorFalso(self._resposta(sql))


class CursorFalso:
    def __init__(self, linhas):
        self._linhas = list(linhas)

    def fetchall(self):
        linhas, self._linhas = self._linhas, []
        return linhas

    def fetchmany(self, n):
        pedaco, self._linhas = self._linhas[:n], self._linhas[n:]
        return pedaco

    def close(self):
        pass


def _titulo(codigo, natureza, valor, categoria="1.01", vencimento="10/03/2025",
            status="Recebido", retencoes=(0, 0, 0, 0, 0, 0), observacao=""):
    """Um título no formato exato que o SELECT do espelho devolve."""
    return (codigo, natureza, valor, categoria, 99, 7, "NF1", "PED1", status,
            vencimento, *retencoes, observacao)


CATALOGO = [("1.01", "Receita de Obras", "Receitas", "3.01", "N", "Receita Bruta")]
CLIENTES = [(99, "CLIENTE TAL LTDA", "12.345.678/0001-90")]
OBRAS = [("D1", "PROJ-A"), ("D2", "PROJ-B")]
CONTAS = [(7, "Bradesco C/C 1234-5")]


def test_titulo_sem_rateio_vai_para_nao_apropriado():
    """Título que ninguém apropriou a uma obra não pode sumir nem ser diluído:
    ele aparece com obra '(não apropriado)', para o erro ficar visível."""
    conn = ConexaoFalsa([_titulo(1, "R", 1000.0)],
                        categorias=CATALOGO, clientes=CLIENTES,
                        obras=OBRAS, contas=CONTAS)
    linhas = list(fato.gerar_linhas_fato(conn))
    assert len(linhas) == 1
    assert dict(zip(fato.COLUNAS_FATO, linhas[0]))["departamento"] == "(não apropriado)"


def test_rateio_divide_o_titulo_entre_as_obras():
    """Um título de 1.000 rateado 700/300 vira duas linhas, e a soma fecha."""
    conn = ConexaoFalsa(
        [_titulo(1, "R", 1000.0)],
        rateios=[(1, "D1", "Obra Um", 70.0, 700.0),
                 (1, "D2", "Obra Dois", 30.0, 300.0)],
        movimentos=[(1, "15/03/2025", "S", 1000.0, 0.0, 0.0, 0.0, 0.0)],
        categorias=CATALOGO, clientes=CLIENTES, obras=OBRAS, contas=CONTAS)
    linhas = [dict(zip(fato.COLUNAS_FATO, l)) for l in fato.gerar_linhas_fato(conn)]

    assert len(linhas) == 2
    assert {l["departamento"] for l in linhas} == {"Obra Um", "Obra Dois"}
    assert {l["projeto"] for l in linhas} == {"PROJ-A", "PROJ-B"}
    assert sum(l["pago_recebido"] for l in linhas) == pytest.approx(1000.0, abs=0.01)


def test_receita_com_imposto_retido_gera_linha_separada():
    """A retenção não é caixa da BWS, mas faz parte do bruto da medição. Por
    isso vira uma LINHA própria: líquida + retida = bruto."""
    conn = ConexaoFalsa(
        [_titulo(1, "R", 1000.0, retencoes=(50.0, 30.0, 0, 0, 0, 0))],
        rateios=[(1, "D1", "Obra Um", 100.0, 1000.0)],
        movimentos=[(1, "15/03/2025", "S", 920.0, 0.0, 0.0, 0.0, 0.0)],
        categorias=CATALOGO, clientes=CLIENTES, obras=OBRAS, contas=CONTAS)
    linhas = [dict(zip(fato.COLUNAS_FATO, l)) for l in fato.gerar_linhas_fato(conn)]

    assert len(linhas) == 2
    retida = next(l for l in linhas if "Retido" in l["categoria"])
    liquida = next(l for l in linhas if "Retido" not in l["categoria"])
    assert retida["pago_recebido"] == pytest.approx(80.0)
    assert liquida["pago_recebido"] == pytest.approx(920.0)
    # bruto = líquido + retido
    assert (liquida["pago_recebido"] + liquida["a_pagar_receber"]
            + retida["pago_recebido"]) == pytest.approx(1000.0)


def test_despesa_entra_negativa():
    """O sinal é o que permite somar receita e despesa na mesma coluna e ler o
    resultado direto. Trocar o sinal dobraria o lucro em vez de zerá-lo."""
    conn = ConexaoFalsa(
        [_titulo(1, "P", 500.0, status="Pago")],
        rateios=[(1, "D1", "Obra Um", 100.0, 500.0)],
        movimentos=[(1, "15/03/2025", "S", 500.0, 0.0, 0.0, 0.0, 0.0)],
        categorias=CATALOGO, clientes=CLIENTES, obras=OBRAS, contas=CONTAS)
    linha = dict(zip(fato.COLUNAS_FATO, next(iter(fato.gerar_linhas_fato(conn)))))
    assert linha["pago_recebido"] == pytest.approx(-500.0)
    assert linha["tipo"] == "2. Contas a Pagar"


def test_titulo_cancelado_nao_entra():
    conn = ConexaoFalsa([_titulo(1, "R", 1000.0, status="CANCELADO")],
                        categorias=CATALOGO, clientes=CLIENTES,
                        obras=OBRAS, contas=CONTAS)
    assert list(fato.gerar_linhas_fato(conn)) == []


def test_titulo_em_aberto_fica_em_a_receber_e_usa_a_data_de_vencimento():
    conn = ConexaoFalsa(
        [_titulo(1, "R", 1000.0, status="A Receber", vencimento="20/12/2025")],
        rateios=[(1, "D1", "Obra Um", 100.0, 1000.0)],
        categorias=CATALOGO, clientes=CLIENTES, obras=OBRAS, contas=CONTAS)
    linha = dict(zip(fato.COLUNAS_FATO, next(iter(fato.gerar_linhas_fato(conn)))))
    assert linha["pago_recebido"] == pytest.approx(0.0)
    assert linha["a_pagar_receber"] == pytest.approx(1000.0)
    assert linha["data"] == dt.date(2025, 12, 20)
    assert linha["situacao_vencimento"] in ("A vencer", "Vencido")


def test_cada_linha_tem_exatamente_as_colunas_da_tabela():
    """Se alguém acrescentar um campo no gerador e esquecer da tabela (ou o
    contrário), o INSERT quebra em produção. Aqui quebra no teste."""
    conn = ConexaoFalsa(
        [_titulo(1, "R", 100.0)],
        rateios=[(1, "D1", "Obra Um", 100.0, 100.0)],
        categorias=CATALOGO, clientes=CLIENTES, obras=OBRAS, contas=CONTAS)
    for linha in fato.gerar_linhas_fato(conn):
        assert len(linha) == len(fato.COLUNAS_FATO)


# ===========================================================================
# 4. Geometria dos gráficos
# ===========================================================================
def test_regua_do_grafico_sempre_inclui_o_zero():
    """Sem o zero na régua, uma barra negativa não teria de onde partir."""
    eixo = graficos.eixo_vertical([120.0, 340.0, 90.0])
    assert eixo["base"] <= 0 <= eixo["topo"]


def test_grafico_vazio_nao_quebra():
    assert graficos.barras_agrupadas([], [("v", "b-caixa", "Caixa")])["vazio"] is True


def test_barra_negativa_desce_a_partir_do_zero():
    g = graficos.barras_agrupadas(
        [{"ano": 2024, "v": -100.0}, {"ano": 2025, "v": 300.0}],
        [("v", "b-despesa", "Despesa")])
    negativa, positiva = g["barras"]
    # no SVG o y cresce para baixo: a barra negativa começa NO zero
    assert negativa["y"] == pytest.approx(g["y_zero"], abs=0.5)
    assert positiva["y"] < g["y_zero"]


def test_com_muitos_meses_os_rotulos_nao_se_amontoam():
    """Seis anos de história são ~70 meses no eixo. Escrever "06/2025" em todos
    vira uma tarja preta: as datas se sobrepõem e não se lê nenhuma."""
    meses = [{"rotulo": f"{m % 12 + 1:02d}/{2020 + m // 12}", "receita": 100.0,
              "despesa": -80.0} for m in range(84)]
    g = graficos.barras_agrupadas(
        meses, [("receita", "b-verde", "Receita"), ("despesa", "b-vermelho", "Despesa")],
        campo_rotulo="rotulo")

    # o desenho tem 810px úteis; com 84 rótulos de ~54px não cabe nem um terço
    assert len(g["rotulos_x"]) < 20
    # e os que sobraram não podem estar colados um no outro
    xs = sorted(r["x"] for r in g["rotulos_x"])
    assert min(b - a for a, b in zip(xs, xs[1:])) >= 40
    # o último mês é o que interessa a quem olha: nunca some
    assert g["rotulos_x"][-1]["texto"] == meses[-1]["rotulo"]


def test_com_poucos_periodos_todos_os_rotulos_aparecem():
    """Pular rótulo onde cabe todo mundo seria esconder informação à toa."""
    anos = [{"ano": a, "receita": 10.0, "despesa": -8.0} for a in range(2019, 2027)]
    g = graficos.barras_agrupadas(
        anos, [("receita", "b-verde", "R"), ("despesa", "b-vermelho", "D")])
    assert len(g["rotulos_x"]) == len(anos)


def test_proporcoes_do_ranking_vao_de_0_a_100():
    itens = graficos.proporcoes([{"valor": -1000.0}, {"valor": -250.0}])
    assert itens[0]["pct"] == 100.0
    assert itens[1]["pct"] == 25.0


# ===========================================================================
# 5. As telas de verdade, pelo HTTP
# ===========================================================================
# Um erro de Jinja (nome trocado, filtro que não existe) só aparece quando a
# página é montada — nenhum dos testes acima o pegaria. Aqui o Flask sobe de
# verdade e as telas são pedidas por HTTP, com o banco dublado.

RESPOSTAS_FALSAS = {
    "DISTINCT TRIM(grupo)": [("Materiais Aplicados",), ("Despesas com Pessoal",)],
    "DISTINCT TRIM(categoria)": [("Salários",), ("Cimento",)],
    "DISTINCT ano": [(2025,), (2024,)],
    "DISTINCT projeto": [("PROJ-A",), ("PROJ-B",)],
    "DISTINCT departamento": [("Obra Um",), ("Obra Dois",)],
    # o carimbo da base: e ele que diz se as listas guardadas ainda valem
    "MAX(fim) FROM execucoes": [(dt.datetime(2026, 9, 2, 3, 12),)],
    "SELECT 1 FROM fato LIMIT 1": [(1,)],
    "FROM execucoes": [("rapida", "agendado",
                        dt.datetime(2026, 9, 2, 3, 0), dt.datetime(2026, 9, 2, 3, 12),
                        True, "185.422 linhas em 11,4 min.", 185422)],
    "COUNT(*) FROM fato": [(185422,)],
}


def _consultar_falso(sql, params=()):
    """Devolve linhas plausíveis para cada consulta, sem banco nenhum.

    A ordem importa: os ramos vão do marcador mais específico para o mais
    genérico. E o final é um erro de propósito — consulta nova sem resposta
    aqui tem de aparecer no teste, não passar batida devolvendo zero.
    """
    for marca, resposta in RESPOSTAS_FALSAS.items():
        if marca in sql:
            return resposta

    # ---- aportes e dividendos (o bloco do fim do DRE) ----
    # Vem antes de tudo: o SQL de aporte cai em vários dos marcadores genéricos
    # lá embaixo, e responder por engano com a linha de outra tela daria um erro
    # de arity difícil de ler.
    if "'Aporte de Parceiro'" in sql:
        if "SUM(-pago_recebido)" in sql:                        # dividendo por obra
            return [("Obra Um", 1200.0)]
        if "ORDER BY 2, 3, 1" in sql:                           # os lançamentos
            return [(dt.date(2025, 2, 3), "Obra Um", "SÓCIO A", "Aporte BWS",
                     "Aportes BWS", 5000.0, "Bradesco C/C", "TED", "capital")]
        if sql.strip().startswith("SELECT COUNT(*) FROM fato"):
            return [(9,)]
        if "= 'Dividendos'" in sql:                             # o quadro à parte
            return [("SÓCIO A", 0.0, 1200.0, 2)]
        if "GROUP BY 1, 2" in sql:                              # por obra / por tipo
            return [("Obra Um", "SÓCIO A", 5000.0, 1000.0, 3),
                    ("Obra Um", "SÓCIO B", 2000.0, 0.0, 1)]
        return [("SÓCIO A", 5000.0, 1000.0, 3),                 # por sócio
                ("SÓCIO B", 2000.0, 0.0, 1)]
    # o resultado por obra do quadro "Resultado x dividendos"; o TRIM só existe
    # nas consultas desse bloco
    if "NULLIF(TRIM(departamento" in sql:
        return [("Obra Um", 6000.0), ("Obra Dois", -800.0)]

    # ---- prestação de contas: as três consultas com bucket "(sem data)" ----
    if "COALESCE(to_char(data" in sql:
        if "TRIM(COALESCE(grupo,'')) = " in sql:                # driver de pessoal
            return [("2025-01", "Obra Um", 900.0)]
        if "departamento = ANY" in sql:                         # despesa admin
            return []
        return [("2025-01", "Obra Um", "PROJ-A", 4000.0, 100.0, -1500.0)]

    # ---- necessidade de caixa ----
    if "emprestimo_tomado" in sql:
        return []
    if "COUNT(*) AS quantas" in sql:                            # obra -> projeto
        return [("Obra Um", "PROJ-A", 12)]
    if "SUM(pago_recebido)" in sql and "GROUP BY 1, 2 " in sql:
        return [(dt.date(2025, 1, 1), "Obra Um", -1000.0),
                (dt.date(2025, 2, 1), "Obra Um", 4000.0)]

    # ---- receita de obra ----
    # O analítico cita `medicao_rotulo` desde que ganhou a coluna da medição, e
    # cairia nos dois ramos abaixo. Vem antes, com um marcador só dele.
    if "situacao_vencimento, pedido_compra" in sql:             # a página de lançamentos
        return [(dt.date(2025, 4, 8), "FORNECEDOR A LTDA", "12.345.678/0001-90",
                 "Despesas com Pessoal", "Salários", "Obra Um", "PROJ-A",
                 "NF77", "folha de março", "Bradesco C/C", "Pago",
                 -3000.0, 0.0, -20.0, -5.0, "",
                 # as quatro que estavam no banco e não apareciam na tela
                 "Quitado", "PC-4471", "Obra Um · Medição 3", 998877)]
    if "medicao_rotulo" in sql and "COUNT(*)" in sql:           # os totais
        return [(3, 7000.0, 300.0, 500.0)]
    if "medicao_rotulo" in sql:                                 # as medições
        return [("OBRA1 | Medição 3", "CLIENTE A", "Obra Um", "PROJ-A",
                 "NF123", "", dt.date(2025, 5, 2), 7000.0, 300.0, 500.0)]
    if "categoria <> 'Receita de Obras'" in sql:
        return [("Estorno de Despesas", 900.0, 0.0, 4)]
    if "FROM fato_recebimentos" in sql:
        return [(dt.date(2025, 5, 2), 7000.0, 0.0, 0.0, 0.0, "Bradesco C/C",
                 "1/1", "credito bancario", "NF123")]

    # ---- despesas analitico: ver o ramo la em cima ----
    if "COUNT(*), SUM(" in sql:                                 # os totais da seleção
        return [(37, -3000.0, -200.0, -25.0)]
    if "DISTINCT COALESCE(NULLIF(grupo" in sql:
        return [("Despesas com Pessoal",), ("Materiais",)]
    if "DISTINCT COALESCE(NULLIF(categoria" in sql:
        return [("Salários",), ("Cimento",)]

    # ---- as telas antigas ----
    if "date_trunc" in sql:                                     # fluxo mensal
        return [(dt.date(2025, 1, 1), 4000.0, -2500.0),
                (dt.date(2025, 2, 1), 6000.0, -3100.0),
                (dt.date(2025, 3, 1), 2000.0, -5200.0)]
    if "GROUP BY ano" in sql:
        return ([(2024, -1000.0), (2025, 2500.0)] if "SUM(pago_recebido)" in sql
                else [(2024, 5000.0, -4000.0), (2025, 9000.0, -6000.0)])
    if "AS retido" in sql:                                      # as linhas do DRE
        # a última coluna é o encargo: juros e multa efetivamente pagos
        return [("1. Contas a Receber", False, "Receita Bruta", 9000.0, 500.0, 0.0),
                ("1. Contas a Receber", True, "Retenções", 300.0, 0.0, 0.0),
                ("2. Contas a Pagar", False, "Despesas com Pessoal", -4000.0, -100.0, -25.0),
                ("2. Contas a Pagar", False, "Materiais", -2000.0, -50.0, 0.0)]
    # o total de juros e multa pagos, que vira uma linha na aba de categorias.
    # Vem antes do ranking porque nao tem HAVING mas soma sobre o mesmo fato.
    if "COALESCE(SUM(CASE WHEN" in sql and "(juros + multa)" in sql:
        return [(-25.0,)]
    if "HAVING" in sql:                                         # ranking de despesas
        return [("Despesas com Pessoal", -4100.0), ("Materiais", -2050.0)]
    if "COUNT(*)" in sql:                                       # maiores credores
        return [("FORNECEDOR A LTDA", -3000.0, -200.0, 12)]
    if "ABS(SUM(" in sql:                                       # comprometido x exec
        return [("Obra Um", -8000.0, -2000.0), ("Obra Dois", -3000.0, -500.0)]
    if "GROUP BY 1" in sql and sql.count("SUM(CASE WHEN tipo") >= 2:
        return [("PROJ-A", 12000.0, -9000.0), ("PROJ-B", 4000.0, -6000.0)]
    if "GROUP BY 1" in sql:                                     # receita por obra
        return [("Obra Um", 7000.0, 300.0, 500.0)]
    if "pago_recebido > 0" in sql:                              # caixa: entra e sai
        return [(18500.0, -12300.0)]
    if sql.count("SUM(CASE WHEN tipo") == 4:                    # resumo do resultado
        return [(9500.0, -6150.0, 9000.0, -6000.0)]
    raise AssertionError(f"consulta sem resposta no dublê: {sql.strip()[:120]}")


CONFIG_FALSA = {
    "projeto_matriz": "PROJ-A", "depto_admin_matriz": "ADM MATRIZ",
    "depto_admin_filial": "ADM FILIAL", "grupo_pessoal": "Despesas com Pessoal",
    "taxa_adm_pct": "1.5", "residual": "1",
}


@pytest.fixture()
def painel(monkeypatch):
    """App do monorepo com o painel ligado e o banco dublado.

    Dublar `consultas.consultar` não basta: `prestacao_dados` tem a própria
    referência para o banco. Um teste que esqueça disso faz o painel abrir a
    conexão de verdade — e como o `.env` da raiz aponta para a produção, é lá
    que ele bateria. A trava em `painel/db.py` recusa isso, mas o certo é o
    dublê cobrir tudo."""
    monkeypatch.setenv("PAINEL_SENHA", "segredo-de-teste")
    from app.apps.painel import consultas, prestacao_dados
    monkeypatch.setattr(consultas, "consultar", _consultar_falso)
    monkeypatch.setattr(prestacao_dados, "config", lambda: dict(CONFIG_FALSA))
    for nome in ("socios", "participacoes", "regras", "ajustes"):
        monkeypatch.setattr(prestacao_dados, nome, lambda *a, **k: [])

    from app.main import create_app
    app = create_app()
    app.config.update(TESTING=True)
    return app.test_client()


def test_sem_senha_nao_ve_nada(painel):
    """São dados financeiros: o padrão é negar. Sem sessão, vai para o login."""
    for caminho in ("/painel/", "/painel/dre", "/painel/configuracoes"):
        r = painel.get(caminho)
        assert r.status_code == 302, caminho
        assert "/painel/entrar" in r.headers["Location"], caminho


def test_senha_errada_nao_entra(painel):
    r = painel.post("/painel/entrar", data={"senha": "chute"})
    assert r.status_code == 401


def test_disparo_da_madrugada_sem_segredo_e_recusado(painel):
    """O endereço que o agendador chama é público por necessidade; quem o
    protege é o segredo do módulo."""
    r = painel.post("/painel/api/sincronizar", json={"modo": "rapida"})
    assert r.status_code == 401
    assert r.get_json()["ok"] is False


def test_visao_geral_abre_e_mostra_os_numeros(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get("/painel/")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "Visão Geral" in html
    assert "Receita líquida" in html
    assert "Geração de caixa" in html
    assert "R$" in html
    assert "<svg" in html          # o gráfico foi desenhado


def test_dre_abre_com_as_tres_leituras(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get("/painel/dre?ano=2025&quebra=grupo")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    # os rótulos são os da tela original, não os que eu tinha inventado
    assert "Receita Bruta de Serviços" in html
    assert "(−) Retenções na fonte" in html
    assert "= Receita Líquida" in html
    assert "= Total Custos/Despesas" in html
    assert "= RESULTADO" in html
    assert "Juros e Multas Pagos" in html      # os encargos voltaram para o DRE
    assert "Fluxo Financeiro" in html          # o gráfico mensal voltou
    assert "Despesas com Pessoal" in html
    assert "Ano: 2025" in html                 # o filtro virou chip na tela


def test_juros_e_multa_pagos_entram_no_total_de_custos(painel):
    """Não basta a linha aparecer na tela: ela tem de SOMAR.

    A primeira versão do DRE deixou os encargos de fora, e o resultado saía
    maior do que é de verdade — um erro que ninguém percebe olhando, porque a
    tela continua parecendo uma planilha certa. O teste ao lado confere que o
    rótulo aparece; este confere a aritmética.
    """
    from app.apps.painel import consultas

    d = consultas.dre_linhas(consultas.Filtros())
    linha = {l["linha"].strip(): l for l in d["linhas"] if l["linha"].strip()}

    juros = linha["Juros e Multas Pagos"]["comprometido"]
    total = linha["= Total Custos/Despesas"]["comprometido"]
    assert round(juros, 2) == -25.00           # veio da última coluna do SQL

    # as linhas recuadas (grupos + encargos) têm de dar exatamente o total
    recuadas = sum(l["comprometido"] for l in d["linhas"]
                   if l["linha"].startswith("  "))
    assert round(recuadas, 2) == round(total, 2)
    assert round(total, 2) == -6175.00         # -6.150 dos grupos, -25 de encargo

    # e o resultado desce do total, encargo incluído
    liquida = linha["= Receita Líquida"]["comprometido"]
    assert round(liquida + total, 2) == round(linha["= RESULTADO"]["comprometido"], 2)


def test_filtros_da_url_chegam_na_consulta(painel, monkeypatch):
    """A tela filtrada tem de poder ser salva nos favoritos e reabrir igual."""
    vistos = []
    from app.apps.painel import consultas
    original = _consultar_falso

    def espiao(sql, params=()):
        vistos.append((sql, list(params)))
        return original(sql, params)

    monkeypatch.setattr(consultas, "consultar", espiao)
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    assert painel.get("/painel/?ano=2025&projeto=PROJ-A&obra=Obra+Um").status_code == 200

    com_filtro = [p for sql, p in vistos if "ano = ANY" in sql]
    assert com_filtro, "nenhuma consulta recebeu o filtro de ano"
    assert [2025] in com_filtro[0]
    assert ["PROJ-A"] in com_filtro[0]
    assert ["Obra Um"] in com_filtro[0]


def test_sair_derruba_a_sessao(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    assert painel.get("/painel/").status_code == 200
    painel.get("/painel/sair")
    assert painel.get("/painel/").status_code == 302


def test_fluxo_de_caixa_abre_e_acumula(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get("/painel/fluxo")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "Fluxo de Caixa" in html
    assert "01/2025" in html and "03/2025" in html
    assert "Caixa acumulado" in html
    # 1500 + 2900 - 3200 = 1200 acumulado no fim
    assert "R$ 1.200,00" in html


def test_resultado_por_obra_troca_o_agrupamento(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    por_projeto = painel.get("/painel/obras").get_data(as_text=True)
    assert "Resultado por Projeto" in por_projeto
    assert "PROJ-A" in por_projeto

    por_obra = painel.get("/painel/obras?nivel=obra").get_data(as_text=True)
    assert "Resultado por Obra" in por_obra


def test_comprometido_vs_executado_calcula_o_percentual(painel):
    """8.000 pagos de 10.000 comprometidos = 80% andado."""
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html = painel.get("/painel/execucao").get_data(as_text=True)
    assert "Comprometido" in html
    assert "80%" in html


def test_lado_a_receber_muda_a_explicacao(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html = painel.get("/painel/execucao?tipo=receber").get_data(as_text=True)
    assert "ainda falta receber" in html


def test_telas_novas_tambem_exigem_login(painel):
    for caminho in ("/painel/fluxo", "/painel/obras", "/painel/execucao"):
        assert painel.get(caminho).status_code == 302, caminho


# ===========================================================================
# 6. As telas que vieram depois
# ===========================================================================
def test_receita_de_obra_abre(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html = painel.get("/painel/receita").get_data(as_text=True)
    assert "Receita de Obra" in html
    assert "OBRA1 | Medição 3" in html
    assert "Retido pelo cliente" in html
    assert "Estorno de Despesas" in html          # as outras receitas


def test_detalhe_da_medicao_abre_os_recebimentos(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html = painel.get("/painel/receita/OBRA1%20%7C%20Medi%C3%A7%C3%A3o%203").get_data(as_text=True)
    assert "02/05/2025" in html
    assert "credito bancario" in html


def test_necessidade_de_caixa_pede_o_conjunto_antes_de_calcular(painel):
    """Sem escolher obra nenhuma não há o que simular — e a tela diz isso em
    vez de mostrar linhas zeradas como se fossem resposta."""
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html = painel.get("/painel/necessidade-caixa").get_data(as_text=True)
    assert "Escolha ao menos uma obra ou projeto" in html


def test_prestacao_avisa_quando_nao_ha_participacao(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html = painel.get("/painel/prestacao").get_data(as_text=True)
    assert "Nenhuma participação cadastrada" in html


def test_parametros_da_prestacao_abrem_em_todas_as_abas(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    for aba, marca in [("socios", "Adicionar ou renomear sócio"),
                       ("participacoes", "Quem participa de qual projeto"),
                       ("regras", "Nova regra"),
                       ("ajustes", "Novo ajuste"),
                       ("geral", "Parâmetros gerais")]:
        html = painel.get(f"/painel/prestacao/parametros?aba={aba}").get_data(as_text=True)
        assert marca in html, aba


def test_a_planilha_e_um_excel_de_verdade(painel):
    """CSV virava oito arquivos soltos. O relatório do dono sempre foi um
    arquivo com várias abas — voltou a ser."""
    from openpyxl import load_workbook

    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get("/painel/baixar/dre")
    assert r.status_code == 200
    assert r.get_data()[:2] == b"PK"                  # é mesmo um .xlsx
    assert ".xlsx" in r.headers["Content-Disposition"]
    assert "attachment" in r.headers["Content-Disposition"]

    folha = load_workbook(io.BytesIO(r.get_data()))["DRE"]
    assert [c.value for c in folha[1]] == ["Linha", "Executado", "Em aberto",
                                           "Comprometido"]
    assert folha.freeze_panes == "A2"                 # cabeçalho fixo


def test_o_relatorio_completo_tem_todas_as_abas(painel):
    """Era assim na tela antiga: um arquivo, uma aba por assunto."""
    from openpyxl import load_workbook

    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get("/painel/baixar/completo")
    assert r.status_code == 200
    livro = load_workbook(io.BytesIO(r.get_data()))
    assert livro.sheetnames == [
        "DRE", "Despesas Categoria", "Top Credores", "Receita de Obra",
        "Outras Receitas", "Despesas Analitico", "Fluxo de Caixa",
        "Resultado por Obra"]


def test_a_aba_de_categorias_fecha_com_a_aba_do_dre(painel):
    """Duas abas do mesmo arquivo não podem mostrar totais diferentes.

    Juros e multa pagos entram no DRE mas não têm categoria no plano financeiro
    do OMIE. A planilha antiga acrescentava a linha de propósito, para as duas
    abas fecharem; sem ela, quem soma a aba de categorias acha que a despesa é
    menor do que o próprio arquivo diz duas abas antes.
    """
    from openpyxl import load_workbook

    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get("/painel/baixar/completo")
    livro = load_workbook(io.BytesIO(r.get_data()))

    categorias = livro["Despesas Categoria"]
    rotulos = [linha[0] for linha in categorias.iter_rows(min_row=2, values_only=True)]
    assert "Juros e Multas Pagos" in rotulos

    soma_categorias = sum(linha[1] for linha in
                          categorias.iter_rows(min_row=2, values_only=True))

    dre = {linha[0].strip(): linha[3] for linha in
           dre_linhas_da_planilha(livro) if linha[0]}
    assert round(soma_categorias, 2) == round(dre["= Total Custos/Despesas"], 2)


def dre_linhas_da_planilha(livro):
    """As linhas da aba DRE, já sem o cabeçalho."""
    return livro["DRE"].iter_rows(min_row=2, values_only=True)


def test_a_planilha_respeita_os_filtros_da_tela(painel, monkeypatch):
    """Baixar tem de trazer o que estava na tela, não a base inteira."""
    vistos = []
    from app.apps.painel import consultas
    original = _consultar_falso

    def espiao(sql, params=()):
        vistos.append(list(params))
        return original(sql, params)

    monkeypatch.setattr(consultas, "consultar", espiao)
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    assert painel.get("/painel/baixar/dre?ano=2025&projeto=PROJ-A").status_code == 200
    assert any([2025] in p for p in vistos)
    assert any(["PROJ-A"] in p for p in vistos)


def test_assunto_de_planilha_desconhecido_responde_404(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    assert painel.get("/painel/baixar/qualquer-coisa").status_code == 404


def test_as_telas_novas_tambem_exigem_login(painel):
    for caminho in ("/painel/receita", "/painel/necessidade-caixa",
                    "/painel/prestacao", "/painel/prestacao/parametros",
                    "/painel/baixar/dre"):
        assert painel.get(caminho).status_code == 302, caminho


def test_importar_prestacao_sem_arquivo_e_recusado(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.post("/painel/api/importar-prestacao")
    assert r.status_code == 400
    assert r.get_json()["ok"] is False


# ---------------------------------------------------------------------------
# As abas de dentro do DRE
# ---------------------------------------------------------------------------
# A primeira conversão jogou fora metade da tela do DRE — Receitas, Top
# Credores e o bloco inteiro de Aportes. O dono viu na hora e reclamou, com
# razão. Estes testes existem para isso não voltar a acontecer sem alguém
# perceber: cada aba tem de abrir e mostrar o que mostrava antes.

def test_as_quatro_abas_do_dre_estao_na_tela(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html = painel.get("/painel/dre").get_data(as_text=True)
    for rotulo in ("Despesas", "Receitas", "Top Credores", "Aportes e dividendos"):
        assert rotulo in html


def test_aba_de_receitas_traz_medicoes_e_outras_receitas(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html = painel.get("/painel/dre?bloco=receitas").get_data(as_text=True)
    assert "Receita de Obra" in html
    assert "Outras Receitas" in html
    assert "OBRA1 | Medição 3" in html
    assert "Estorno de Despesas" in html
    assert "Bruto faturado" in html


def test_aba_de_credores_lista_e_leva_ao_analitico(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html = painel.get("/painel/dre?bloco=credores").get_data(as_text=True)
    assert "FORNECEDOR A LTDA" in html
    # o nome tem de ser clicável: é assim que se chega ao lançamento
    assert "credor=FORNECEDOR" in html.replace("+", " ").replace("%20", " ")


def test_aba_de_aportes_mostra_os_quatro_recortes(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html = painel.get("/painel/dre?bloco=aportes").get_data(as_text=True)
    assert "Aportes e devoluções" in html
    assert "Por sócio ou parceiro" in html
    assert "Por obra" in html
    assert "Por tipo" in html
    assert "Lançamentos" in html
    assert "Falta p/ igualar" in html
    assert "SÓCIO A" in html


def test_aporte_nao_entra_em_nenhuma_linha_do_dre(painel):
    """A regra que mais importa aqui: aporte é fluxo, não resultado.

    Se um dia alguém somar aporte na despesa, o total do DRE muda sem que
    nenhuma tela reclame — e o dono decide errado com base nisso."""
    from app.apps.painel import consultas
    where, _ = consultas.Filtros().where("analise = 'DRE'")
    sql = f"SELECT 1 FROM fato{where}"
    assert "Aporte" not in sql


def test_dividendo_fica_fora_do_saldo_de_aporte(painel):
    """Dividendo é distribuição de lucro. Abatê-lo do saldo faria parecer que o
    sócio retirou o capital que colocou."""
    from app.apps.painel import consultas
    assert "Dividendos" not in consultas.NO_SALDO


def test_resultado_x_dividendos_calcula_o_disponivel(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html = painel.get("/painel/dre?bloco=aportes").get_data(as_text=True)
    assert "Resultado × dividendos" in html
    assert "Disponível" in html
    # 6.000,00 − 800,00 de resultado, menos 1.200,00 de dividendo pago
    assert "4.000,00" in html


def test_a_aba_pedida_sobrevive_a_mudanca_de_filtro(painel):
    """Trocar a quebra das despesas não pode jogar a pessoa de volta na
    primeira aba, nem apagar o filtro de obra que ela acabou de montar."""
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html = painel.get("/painel/dre?bloco=despesas&ano=2025&quebra=categoria"
                      ).get_data(as_text=True)
    assert '<input type="hidden" name="bloco" value="despesas">' in html
    assert '<input type="hidden" name="ano" value="2025">' in html


def test_aba_desconhecida_cai_na_primeira_em_vez_de_quebrar(painel):
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get("/painel/dre?bloco=inventado")
    assert r.status_code == 200
    assert "Despesas por Grupo" in r.get_data(as_text=True)


def test_o_excel_de_aportes_tem_uma_aba_por_recorte(painel):
    from openpyxl import load_workbook

    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get("/painel/baixar/aportes")
    assert r.status_code == 200
    livro = load_workbook(io.BytesIO(r.get_data()))
    assert livro.sheetnames == [
        "Aportes por Socio", "Aportes por Obra", "Aportes por Tipo",
        "Dividendos", "Lancamentos de Aporte", "Resultado x Dividendos"]


# ===========================================================================
# 8. O filtro que não pegava — e a varredura da classe inteira
# ===========================================================================
# Um formulário GET que manda o MESMO nome duas vezes — uma escondida com o
# valor velho, outra na caixa com o valor novo — vira "?visao=comprometido&
# visao=aberto". O Flask lê o PRIMEIRO. Efeito na tela: a pessoa escolhe, clica
# em Aplicar, e o filtro volta ao que era, sem erro nenhum.
#
# Aconteceu em duas telas ao mesmo tempo (Analítico e DRE), porque o jeito de
# levar os filtros adiante é o mesmo nas duas. Por isso o teste não confere um
# caso: varre TODOS os formulários de TODAS as telas.
import html.parser


class _Formularios(html.parser.HTMLParser):
    """Coleta, por formulário, os campos escondidos e os que a pessoa preenche."""

    def __init__(self):
        super().__init__()
        self.formularios = []
        self._atual = None

    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        if tag == "form":
            self._atual = {"metodo": (a.get("method") or "get").lower(),
                           "ocultos": [], "visiveis": []}
            self.formularios.append(self._atual)
            return
        if self._atual is None or "name" not in a:
            return
        if tag == "input" and (a.get("type") or "").lower() == "hidden":
            self._atual["ocultos"].append(a["name"])
        elif tag in ("input", "select", "textarea"):
            self._atual["visiveis"].append(a["name"])

    def handle_endtag(self, tag):
        if tag == "form":
            self._atual = None


TELAS_COM_FILTRO = [
    "/painel/", "/painel/dre", "/painel/analitico", "/painel/receita",
    "/painel/fluxo", "/painel/obras", "/painel/execucao",
]


@pytest.mark.parametrize("caminho", TELAS_COM_FILTRO)
def test_nenhum_formulario_manda_o_mesmo_campo_duas_vezes(painel, caminho):
    """O valor velho não pode viajar junto com o novo.

    Se um nome aparece escondido E numa caixa do mesmo formulário, o servidor
    recebe os dois e fica com o primeiro — o filtro simplesmente não pega, e
    não há erro nenhum para investigar.
    """
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get(caminho)
    assert r.status_code == 200

    leitor = _Formularios()
    leitor.feed(r.get_data(as_text=True))
    for i, form in enumerate(leitor.formularios):
        if form["metodo"] != "get":
            continue
        repetidos = sorted(set(form["ocultos"]) & set(form["visiveis"]))
        assert not repetidos, (
            f"{caminho}: o formulário {i} manda {repetidos} escondido e na tela "
            f"ao mesmo tempo — o valor velho vence e o filtro não pega")


# ===========================================================================
# 9. O que veio do uso real: velocidade, filtros que ficam, faixa de data
# ===========================================================================
def test_trocar_de_aba_nao_joga_o_filtro_fora(painel):
    """Estar numa obra no DRE e clicar em "Despesas Analítico" tem de continuar
    naquela obra. Antes as abas do topo iam para a tela limpa, e quem tinha
    acabado de montar o filtro montava tudo de novo."""
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get("/painel/dre?ano=2025&obra=Obra+Um&trf=1")
    html_ = r.get_data(as_text=True)

    # o link da aba do Analítico tem de carregar os três filtros
    import re as _re
    destinos = _re.findall(r'class="topo-aba[^"]*"\s+href="([^"]+)"', html_)
    analitico = [d for d in destinos if "/analitico" in d]
    assert analitico, "a aba do Analítico sumiu do topo"
    assert "ano=2025" in analitico[0]
    assert "obra=Obra+Um" in analitico[0] or "obra=Obra%20Um" in analitico[0]
    assert "trf=1" in analitico[0]


def test_a_faixa_de_data_chega_na_consulta(painel, monkeypatch):
    """O filtro de faixa não pode ficar só bonito na tela."""
    from app.apps.painel import consultas
    vistos = []
    original = _consultar_falso

    def espiao(sql, params=()):
        vistos.append((sql, list(params)))
        return original(sql, params)

    monkeypatch.setattr(consultas, "consultar", espiao)
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get("/painel/analitico?de=2025-01-01&ate=2025-03-31")
    assert r.status_code == 200

    da_pagina = [(s, p) for s, p in vistos if "situacao_vencimento, pedido_compra" in s]
    assert da_pagina, "a consulta dos lançamentos não rodou"
    sql, params = da_pagina[0]
    assert "data >= CAST(" in sql and "data <= CAST(" in sql
    assert "2025-01-01" in params and "2025-03-31" in params
    # e a tela mostra que está filtrada, senão o número parece o total
    assert "01/01/2025" in r.get_data(as_text=True)


def test_faixa_invertida_e_endireitada_em_vez_de_nao_trazer_nada(painel, monkeypatch):
    """Digitar o fim antes do começo devolveria zero lançamentos sem dizer por
    quê. Vira a faixa e segue."""
    from app.apps.painel import consultas
    vistos = []
    monkeypatch.setattr(consultas, "consultar",
                        lambda s, p=(): (vistos.append((s, list(p))),
                                         _consultar_falso(s, p))[1])
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    painel.get("/painel/analitico?de=2025-03-31&ate=2025-01-01")
    sql, params = [x for x in vistos if "situacao_vencimento, pedido_compra" in x[0]][0]
    assert params.index("2025-01-01") < params.index("2025-03-31")


def test_data_invalida_na_barra_de_endereco_nao_derruba_a_tela(painel):
    """A faixa é conveniência: texto estranho é ignorado, não vira erro 500."""
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get("/painel/analitico?de=ontem&ate=');DROP+TABLE")
    assert r.status_code == 200


def test_o_analitico_mostra_o_que_estava_escondido_no_banco(painel):
    """Situação do vencimento, pedido de compra e medição já existiam na base e
    não apareciam em tela nenhuma."""
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html_ = painel.get("/painel/analitico").get_data(as_text=True)
    assert "Quitado" in html_                    # situação do vencimento
    assert "PC-4471" in html_                    # pedido de compra
    assert "Medição 3" in html_                  # a medição em que o custo caiu


def test_a_planilha_do_analitico_leva_as_colunas_novas(painel):
    from openpyxl import load_workbook

    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    r = painel.get("/painel/baixar/analitico")
    aba = load_workbook(io.BytesIO(r.get_data()))["Despesas Analitico"]
    cabecalho = [c.value for c in next(aba.iter_rows(max_row=1))]
    for coluna in ("Situação do vencimento", "Pedido de compra", "Medição",
                   "Nº no OMIE"):
        assert coluna in cabecalho


def test_saber_se_a_base_esta_vazia_nao_conta_a_base_inteira(painel, monkeypatch):
    """`COUNT(*)` numa tabela de 185 mil linhas para responder "existe alguma?"
    era uma varredura completa em TODA abertura de TODA tela."""
    from app.apps.painel import consultas
    vistos = []
    monkeypatch.setattr(consultas, "consultar",
                        lambda s, p=(): (vistos.append(s), _consultar_falso(s, p))[1])
    consultas.base_vazia()
    assert vistos == ["SELECT 1 FROM fato LIMIT 1"]


def test_as_listas_de_opcoes_nao_sao_refeitas_a_cada_tela(painel, monkeypatch):
    """Anos, projetos e obras só mudam quando entra carga nova. Refazer as três
    varreduras a cada clique era o grosso do tempo de abertura."""
    from app.apps.painel import consultas
    consultas.esquecer_listas()
    contadas = []
    monkeypatch.setattr(consultas, "consultar",
                        lambda s, p=(): (contadas.append(s), _consultar_falso(s, p))[1])

    consultas.opcoes_de_filtro()
    distintos = lambda: sum(1 for s in contadas if "SELECT DISTINCT" in s)
    assert distintos() == 3                      # a primeira vez paga
    consultas.opcoes_de_filtro()
    consultas.opcoes_de_filtro()
    assert distintos() == 3                      # as seguintes, não


def test_carga_nova_joga_fora_a_lista_guardada(painel, monkeypatch):
    """O contrário seria pior que a lentidão: obra nova entra na base e não
    aparece no filtro até o servidor reiniciar."""
    from app.apps.painel import consultas
    consultas.esquecer_listas()
    carimbo = ["2026-09-02 03:12:00"]

    def falso(sql, params=()):
        if "MAX(fim) FROM execucoes" in sql:
            return [(carimbo[0],)]
        if "DISTINCT departamento" in sql:
            return [("Obra Um",)] if carimbo[0].endswith("12:00") else [
                ("Obra Um",), ("Obra Nova",)]
        return _consultar_falso(sql, params)

    monkeypatch.setattr(consultas, "consultar", falso)
    assert consultas.opcoes_de_filtro()["obras"] == ["Obra Um"]
    carimbo[0] = "2026-09-03 04:00:00"           # terminou uma carga
    assert consultas.opcoes_de_filtro()["obras"] == ["Obra Um", "Obra Nova"]


def test_a_tela_diz_quanto_tempo_levou(painel):
    """"Está lento" precisa de um número, senão otimizar é adivinhar."""
    painel.post("/painel/entrar", data={"senha": "segredo-de-teste"})
    html_ = painel.get("/painel/analitico").get_data(as_text=True)
    assert "Tela montada em" in html_
    assert "consulta" in html_ and "ao banco" in html_
