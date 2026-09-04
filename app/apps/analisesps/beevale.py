# -*- coding: utf-8 -*-
"""
As planilhas do BeeVale — o cartão de benefício dos terceirizados.

São DUAS COISAS DIFERENTES, e confundi-las é o erro que dá trabalho desfazer:

  1. CADASTRO — pessoas que ainda não têm cartão. Cola-se a lista de e-mails
     ou CPFs, e sai uma planilha de cadastro para subir no portal do BeeVale.
     Não escreve em lugar nenhum: lê a planilha "Dados Documentos" e devolve
     um arquivo para baixar. É reversível por natureza — se sair errado,
     basta gerar de novo.

  2. GERAR — cards de pagamento que já existem no Pipefy. Para cada card:
     busca CPF e valor NO CARD, monta o par de planilhas (Pagamento e
     Cadastro), sobe as duas no Drive e ESCREVE DE VOLTA no card os links e a
     Documentação Fiscal. **Este segundo não tem desfazer.**

Tudo isto veio de um Apps Script (`gerarPlanilhasBeeVale` e
`gerarCadastroBeeValePorLinhasErro`), passou pelo Streamlit e chega aqui com o
mesmo layout de coluna — porque quem recebe o arquivo é o portal do BeeVale,
que não sabe que o programa mudou. Não mexa na ordem nem nos títulos das
colunas sem conferir lá.
"""
from __future__ import annotations

import io
import logging
import os
import re

logger = logging.getLogger("analisesps.beevale")

# A planilha com os dados de cadastro dos colaboradores.
PLANILHA_DADOS = os.getenv(
    "ANALISESPS_SHEET_DADOS_DOCUMENTOS",
    "1fqi4QUOVGUd1_4Gg4vK5qP_IMOSgFaw8DD9MDgmM3vo")
ABA_DADOS = "Dados Documentos"

# O domínio que vira o "e-mail" de cada colaborador no portal do BeeVale: o
# CPF sem pontuação na frente. Não é caixa de correio de verdade — é como o
# portal identifica a pessoa.
DOMINIO = "@bwsconstrucoes.com.br"

# O tipo do arquivo. Mesmo valor do `drive.py`, repetido aqui para a resposta
# de download não precisar importar o Drive só por causa de uma constante.
MIME_XLSX = ("application/vnd.openxmlformats-officedocument"
             ".spreadsheetml.sheet")

# Teto de cards por geração. Cada card são duas subidas no Drive e uma escrita
# no Pipefy; passar disso é pedir para estourar o tempo da requisição no meio,
# com metade dos cards alterados e metade não.
MAXIMO_POR_VEZ = 40


class ErroDoBeeVale(RuntimeError):
    """Falha com mensagem já pronta para a tela."""


def pasta_do_drive() -> str:
    """A pasta onde os arquivos são guardados. Vazio = não configurada."""
    from . import credenciais
    return credenciais.token("DRIVE_FOLDER_ID", "").strip()


# ---------------------------------------------------------------------------
# Arrumação dos dados — as mesmas regras do Apps Script, sem invenção
# ---------------------------------------------------------------------------
def so_digitos(valor) -> str:
    return re.sub(r"\D+", "", str(valor or ""))


def extrair_cpfs(texto: str) -> list:
    """Os CPFs de um texto colado, na ordem, sem repetir.

    Procura primeiro os e-mails no formato `<11 dígitos>@bwsconstrucoes...`,
    que é como o portal do BeeVale devolve a lista de erros. Só se não achar
    nenhum é que cai para números soltos de 11 dígitos — a ordem importa,
    porque um texto de erro costuma ter os dois, e o e-mail é o confiável."""
    texto = str(texto or "")
    achados = re.findall(r"(\d{11})" + re.escape(DOMINIO), texto, flags=re.I)
    if not achados:
        achados = re.findall(r"\b\d{11}\b", texto)
    return list(dict.fromkeys(achados))


def normaliza_cpf(valor) -> str:
    """Onze dígitos, com os zeros à esquerda de volta.

    A planilha guarda CPF como número, e o Google entrega "1234567890" para um
    CPF que começa com zero — e às vezes "1234567890.0". Sem repor o zero, o
    CPF nunca casa."""
    if valor in (None, ""):
        return ""
    texto = str(valor).strip()
    if re.fullmatch(r"\d+(\.0+)?", texto):
        texto = re.sub(r"\.0+$", "", texto)
    digitos = so_digitos(texto)
    if not digitos:
        return ""
    return digitos.rjust(11, "0")[-11:]


def formata_cpf(valor) -> str:
    digitos = so_digitos(valor)
    if len(digitos) != 11:
        return str(valor or "")
    return f"{digitos[0:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:11]}"


def nome_impresso(nome: str) -> str:
    """O que cabe no cartão: primeiro nome e último sobrenome."""
    partes = [p for p in str(nome or "").strip().split() if p]
    if not partes:
        return ""
    if len(partes) == 1:
        return partes[0]
    return f"{partes[0]} {partes[-1]}"


def data_do_cadastro(valor) -> str:
    """Data de nascimento em DD/MM/AAAA, venha como vier da planilha."""
    texto = str(valor or "").strip()
    if not texto:
        return ""
    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", texto):
        return texto
    achado = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", texto)
    if achado:
        return f"{achado.group(3)}/{achado.group(2)}/{achado.group(1)}"
    achado = re.fullmatch(r"(\d{2})-(\d{2})-(\d{4})", texto)
    if achado:
        return f"{achado.group(1)}/{achado.group(2)}/{achado.group(3)}"
    return texto


def telefone_br(valor) -> str:
    """Celular como (DD) 90000-0000, tirando o 55 do país se vier junto."""
    digitos = so_digitos(valor)
    if not digitos:
        return ""
    if len(digitos) >= 12 and digitos.startswith("55"):
        digitos = digitos[2:]
    if len(digitos) > 11:
        digitos = digitos[-11:]
    if len(digitos) == 11:
        return f"({digitos[0:2]}) {digitos[2:7]}-{digitos[7:]}"
    if len(digitos) == 10:
        return f"({digitos[0:2]}) {digitos[2:6]}-{digitos[6:]}"
    return str(valor or "")


def registro(nome, nascimento, telefone, cpf) -> dict:
    """Uma linha do cadastro, do jeito que o portal do BeeVale espera."""
    return {
        "nome_completo": str(nome or "").strip(),
        "nome_impresso": nome_impresso(nome),
        "cpf": formata_cpf(cpf),
        "email": so_digitos(cpf) + DOMINIO,
        "nascimento": data_do_cadastro(nascimento),
        "celular": telefone_br(telefone),
    }


def valor_do_card(texto) -> float:
    """"1.234,56" -> 1234.56. O que não der, vira zero — nunca estoura."""
    texto = str(texto or "").strip()
    if not texto:
        return 0.0
    try:
        return float(texto.replace(".", "").replace(",", "."))
    except ValueError:
        return 0.0


# ---------------------------------------------------------------------------
# A planilha "Dados Documentos"
# ---------------------------------------------------------------------------
def _indice(cabecalho: list, *nomes) -> int | None:
    """A posição da coluna, aceitando os apelidos que a planilha já teve."""
    achatado = [str(c).strip().lower() for c in cabecalho]
    for nome in nomes:
        alvo = nome.strip().lower()
        if alvo in achatado:
            return achatado.index(alvo)
    return None


def buscar_por_cpf(cpfs: list) -> tuple[list, list]:
    """Os cadastros dos CPFs pedidos. Devolve (encontrados, não encontrados).

    Lê a aba inteira de uma vez, e aqui isso é aceitável: são centenas de
    linhas, não as 59 mil da SPsBD. Se um dia crescer, esta é a leitura a
    trocar (ver a regra 1 do HISTORICO)."""
    from .credenciais import cliente, com_retry

    aba = com_retry(lambda: cliente().open_by_key(PLANILHA_DADOS)
                    .worksheet(ABA_DADOS))
    linhas = com_retry(aba.get_all_values)
    if not linhas:
        return [], list(cpfs)

    cabecalho = linhas[0]
    i_nome = _indice(cabecalho, "Nome Completo")
    i_nascimento = _indice(cabecalho, "Data de Nascimento")
    i_telefone = _indice(cabecalho, "Telefone Celular", "Telefone", "Celular")
    i_cpf_num = _indice(cabecalho, "CPF Números")
    i_cpf = _indice(cabecalho, "CPF (Cadastro de Pessoa Física)", "CPF")

    faltando = [rotulo for rotulo, indice in (
        ("Nome Completo", i_nome), ("Data de Nascimento", i_nascimento),
        ("Telefone", i_telefone),
        ("CPF", i_cpf_num if i_cpf_num is not None else i_cpf),
    ) if indice is None]
    if faltando:
        raise ErroDoBeeVale(
            "A planilha \"Dados Documentos\" está sem estas colunas: "
            + ", ".join(faltando)
            + ". Alguém renomeou o cabeçalho — conserte lá, não aqui.")

    def celula(linha, indice):
        if indice is None or indice >= len(linha):
            return ""
        return str(linha[indice]).strip()

    procurados = set(cpfs)
    achados = {}
    for linha in linhas[1:]:
        cpf = (normaliza_cpf(celula(linha, i_cpf_num))
               or normaliza_cpf(celula(linha, i_cpf)))
        if not cpf or cpf not in procurados or cpf in achados:
            continue
        achados[cpf] = registro(celula(linha, i_nome),
                                celula(linha, i_nascimento),
                                celula(linha, i_telefone), cpf)

    return ([achados[c] for c in cpfs if c in achados],
            [c for c in cpfs if c not in achados])


# ---------------------------------------------------------------------------
# Os dois arquivos
#
# O layout das colunas é contrato com o portal do BeeVale — não é escolha de
# estilo. Tudo vai como TEXTO menos o valor e os dias úteis: um CPF que o
# Excel entenda como número perde o zero da frente, e o portal recusa.
# ---------------------------------------------------------------------------
COLUNAS_CADASTRO = ["Nome completo", "Nome Impresso no Cartão", "CPF",
                    "Email", "Data de nascimento", "Celular"]

COLUNAS_PAGAMENTO = ["Nome", "Email", "Carteira", "Benefício", "Valor",
                     "Tipo de Recarga", "Dias úteis",
                     "Documento de Identificação",
                     "Nome Completo (Colaborador)", "Centro de Custo",
                     "Categoria"]


def _fechar(planilha) -> bytes:
    memoria = io.BytesIO()
    planilha.save(memoria)
    return memoria.getvalue()


def cadastro_xlsx(registros: list) -> bytes:
    """A planilha de cadastro, para subir no portal do BeeVale."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    planilha = Workbook()
    aba = planilha.active
    aba.title = "Cadastro BeeVale"
    aba.append(COLUNAS_CADASTRO)
    for r in registros:
        aba.append([r.get("nome_completo", ""), r.get("nome_impresso", ""),
                    r.get("cpf", ""), r.get("email", ""),
                    r.get("nascimento", ""), r.get("celular", "")])
    for coluna in range(1, len(COLUNAS_CADASTRO) + 1):
        letra = get_column_letter(coluna)
        for celula in aba[letra]:
            celula.number_format = "@"
        aba.column_dimensions[letra].width = 24
    aba.freeze_panes = "A2"
    return _fechar(planilha)


def pagamento_xlsx(linhas: list) -> bytes:
    """A planilha de recarga. `linhas`: [{nome, email, valor, cpf, card}]."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    planilha = Workbook()
    aba = planilha.active
    aba.title = "Pagamento BeeVale"
    aba.append(COLUNAS_PAGAMENTO)
    for linha in linhas:
        aba.append([linha.get("nome", ""), linha.get("email", ""),
                    "Produção", "Livre", float(linha.get("valor", 0) or 0),
                    "Mensal", 0, linha.get("cpf", ""), linha.get("nome", ""),
                    str(linha.get("card", "")), "Terceirizados"])
    for coluna in (1, 2, 3, 4, 6, 8, 9, 10, 11):        # texto
        for celula in aba[get_column_letter(coluna)]:
            celula.number_format = "@"
    for celula in aba[get_column_letter(5)]:            # Valor
        celula.number_format = "0.00"
    for celula in aba[get_column_letter(7)]:            # Dias úteis
        celula.number_format = "0"
    aba.freeze_panes = "A2"
    return _fechar(planilha)


# ---------------------------------------------------------------------------
# A descrição do card
# ---------------------------------------------------------------------------
_LINHAS_DE_LINK = re.compile(
    r"(?i)^\s*(Pagamento BeeVale|Cadastro BeeVale|Planilha BeeVale)\s*:")


def montar_descricao(descricao_atual, link_pagamento: str,
                     link_cadastro: str) -> str:
    """Põe os links no topo, PRESERVANDO o resto da descrição.

    Só as linhas de link antigas saem — e saem para não empilhar uma leva
    nova a cada geração. O texto que a pessoa escreveu fica: é o histórico de
    quem pediu o quê, e este é o único lugar do módulo que poderia apagá-lo."""
    mantidas = [linha for linha in str(descricao_atual or "").strip().split("\n")
                if not _LINHAS_DE_LINK.match(linha)]
    resto = "\n".join(mantidas).strip()
    topo = (f"Pagamento BeeVale: {link_pagamento}\n"
            f"Cadastro BeeVale: {link_cadastro}")
    return f"{topo}\n\n{resto}" if resto else topo


# ---------------------------------------------------------------------------
# A geração completa
# ---------------------------------------------------------------------------
def preparar(ids: list) -> dict:
    """Busca no Pipefy tudo o que os cards precisam, SEM escrever nada.

    Separado de propósito da escrita: dá para mostrar ao operador o que vai
    acontecer — e o que já está impedido — antes de qualquer ação sem volta.
    Devolve {'prontos': [...], 'erros': [...]}."""
    from . import pipefy

    ids = [str(i).strip() for i in ids if str(i).strip()]
    if not ids:
        raise ErroDoBeeVale("Nenhuma SP selecionada.")
    if len(ids) > MAXIMO_POR_VEZ:
        raise ErroDoBeeVale(
            f"São no máximo {MAXIMO_POR_VEZ} SPs por vez — cada uma são duas "
            "subidas no Drive e uma escrita no Pipefy, e passar disso é pedir "
            "para parar no meio. Faça em levas.")

    cards = pipefy.buscar_cards(ids)

    erros, cpf_do_card = [], {}
    for sp_id in ids:
        card = cards.get(sp_id)
        if not card:
            erros.append({"sp": sp_id,
                          "motivo": "O Pipefy não devolveu este card."})
            continue
        cpf = pipefy.extrair_cpf(card["campos"].get(pipefy.CAMPO_CADASTRO, ""))
        if not cpf:
            erros.append({"sp": sp_id,
                          "motivo": "O campo \"Cadastro BeeVale\" do card está "
                                    "vazio ou não tem um CPF."})
            continue
        cpf_do_card[sp_id] = cpf

    cadastros = (pipefy.buscar_cadastros(list(dict.fromkeys(cpf_do_card.values())))
                 if cpf_do_card else {})

    prontos = []
    for sp_id in ids:
        cpf = cpf_do_card.get(sp_id)
        if not cpf:
            continue
        cadastro = cadastros.get(cpf)
        if not cadastro:
            erros.append({"sp": sp_id,
                          "motivo": f"O CPF {cpf} não está na database BeeVale "
                                    "do Pipefy."})
            continue
        nome = (cadastro.get("nome_completo") or "").strip()
        if not nome:
            erros.append({"sp": sp_id,
                          "motivo": f"O cadastro do CPF {cpf} está sem Nome "
                                    "Completo."})
            continue
        prontos.append({
            "sp": sp_id, "cpf": (cadastro.get("cpf") or cpf).strip(),
            "nome": nome,
            "valor": valor_do_card(cards[sp_id]["campos"].get(pipefy.CAMPO_VALOR, "")),
            "cadastro": cadastro,
            "descricao_atual": cards[sp_id]["campos"].get(pipefy.CAMPO_DESCRICAO, ""),
        })
    return {"prontos": prontos, "erros": erros}


def arquivos_do_card(pronto: dict) -> tuple[bytes, bytes]:
    """O par de arquivos de um card: (pagamento, cadastro)."""
    cadastro = pronto["cadastro"]
    email = so_digitos(pronto["cpf"]) + DOMINIO
    pagamento = pagamento_xlsx([{
        "nome": pronto["nome"], "email": email, "valor": pronto["valor"],
        "cpf": formata_cpf(pronto["cpf"]), "card": pronto["sp"]}])
    ficha = cadastro_xlsx([registro(
        pronto["nome"], cadastro.get("data_de_nascimento", ""),
        cadastro.get("telefone_celular", ""), pronto["cpf"])])
    return pagamento, ficha


def nomes_dos_arquivos(sp_id: str) -> tuple[str, str]:
    from .horario import agora
    dia = agora().strftime("%d.%m.%Y")
    return (f"{sp_id}_{dia}.pagamento.beevale.xlsx",
            f"{sp_id}_{dia}.cadastro.beevale.xlsx")


def gerar(ids: list, escrever_no_pipefy: bool = True) -> dict:
    """O fluxo inteiro: monta, sobe no Drive e escreve de volta nos cards.

    ⚠️ COM `escrever_no_pipefy=True` ISTO NÃO TEM DESFAZER — a descrição do
    card é reescrita e a Documentação Fiscal vira "BeeVale".

    A ordem é deliberada: **primeiro tudo o que pode falhar sem estragar**
    (buscar, montar, subir), e só no fim a escrita nos cards. Se o Drive
    recusar, nenhum card foi tocado; se a escrita falhar, os arquivos já
    existem e o operador recebe os links para colar na mão. O contrário —
    marcar o card e depois descobrir que o arquivo não subiu — deixaria um
    card dizendo que está pronto quando não está.

    Devolve {'feitos', 'erros', 'atualizados', 'nao_atualizados'}."""
    from . import drive

    pasta = pasta_do_drive()
    if not pasta:
        raise ErroDoBeeVale(
            "A pasta do Drive não está configurada. Defina DRIVE_FOLDER_ID "
            "no Render (ou na aba Credenciais) com o identificador da pasta — "
            "e ela precisa ser de um Drive Compartilhado. Enquanto isso não "
            "estiver feito, a geração não sai do lugar.")

    preparado = preparar(ids)
    erros = list(preparado["erros"])
    feitos, a_escrever = [], []

    for pronto in preparado["prontos"]:
        try:
            pagamento, ficha = arquivos_do_card(pronto)
            nome_pag, nome_cad = nomes_dos_arquivos(pronto["sp"])
            subiu_pag = drive.subir_xlsx(pagamento, nome_pag, pasta)
            subiu_cad = drive.subir_xlsx(ficha, nome_cad, pasta)
        except Exception as e:  # noqa: BLE001 — um card ruim não para os outros
            logger.exception("Análise de SPs: falhou gerar o BeeVale da SP %s",
                             pronto["sp"])
            erros.append({"sp": pronto["sp"], "motivo": str(e)})
            continue

        feitos.append({"sp": pronto["sp"], "nome": pronto["nome"],
                       "valor": pronto["valor"],
                       "pagamento": subiu_pag["link"],
                       "cadastro": subiu_cad["link"]})
        a_escrever.append({
            "card": pronto["sp"],
            "descricao": montar_descricao(pronto["descricao_atual"],
                                          subiu_pag["link"], subiu_cad["link"])})

    atualizados, nao_atualizados = [], []
    if a_escrever and escrever_no_pipefy:
        from . import pipefy
        try:
            falhas = pipefy.atualizar_descricao_e_doc_fiscal(a_escrever)
        except Exception as e:  # noqa: BLE001
            logger.exception("Análise de SPs: falhou escrever nos cards")
            falhas = [item["card"] for item in a_escrever]
            erros.append({"sp": "(todos)",
                          "motivo": f"Os arquivos subiram no Drive, mas a "
                                    f"escrita nos cards falhou: {e}"})
        nao_atualizados = list(falhas)
        atualizados = [item["card"] for item in a_escrever
                       if item["card"] not in falhas]
        for card in falhas:
            erros.append({"sp": card,
                          "motivo": "Os arquivos subiram no Drive, mas não "
                                    "consegui atualizar a descrição e a "
                                    "Documentação Fiscal do card. Os links "
                                    "estão acima — dá para colar na mão."})
    elif a_escrever:
        nao_atualizados = [item["card"] for item in a_escrever]

    logger.info("Análise de SPs: BeeVale — %d gerado(s), %d card(s) "
                "atualizado(s), %d erro(s).",
                len(feitos), len(atualizados), len(erros))
    return {"feitos": feitos, "erros": erros, "atualizados": atualizados,
            "nao_atualizados": nao_atualizados}
