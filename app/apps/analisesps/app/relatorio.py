# -*- coding: utf-8 -*-
"""
relatorio.py — agregações e exportação (XLSX/PDF) da aba Relatório.

Funções puras (não dependem do Streamlit) para facilitar teste:
  - agregar(df, dim)         -> DataFrame [dim, Qtd, Total]
  - top_credores(df, n)      -> DataFrame [CPF/CNPJ, Credor, Qtd, Total]
  - aging_vencidos(df)       -> DataFrame [Faixa, Qtd, Total]
  - kpis(df)                 -> dict
  - gerar_xlsx(...)          -> bytes (.xlsx)
  - gerar_pdf(...)           -> bytes (.pdf)
"""

import io
import pandas as pd

VAZIO = "(vazio)"


def _moeda(v) -> str:
    try:
        n = float(v or 0)
    except (TypeError, ValueError):
        n = 0.0
    s = f"{abs(n):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ ({s})" if n < 0 else f"R$ {s}"


def agregar(df: pd.DataFrame, dim_col: str, label_vazio: str = VAZIO) -> pd.DataFrame:
    cols = [dim_col, "Qtd", "Total"]
    if df.empty or dim_col not in df.columns:
        return pd.DataFrame(columns=cols)
    g = df.assign(_d=df[dim_col].astype(str).str.strip().replace("", label_vazio))
    out = (g.groupby("_d").agg(Qtd=("id", "count"), Total=("valor_num", "sum"))
           .reset_index().rename(columns={"_d": dim_col})
           .sort_values("Total", ascending=False).reset_index(drop=True))
    return out


def top_credores(df: pd.DataFrame, n: int = 15) -> pd.DataFrame:
    cols = ["CPF/CNPJ", "Credor", "Qtd", "Total"]
    if df.empty:
        return pd.DataFrame(columns=cols)

    def _maiscomum(s):
        m = s.astype(str).mode()
        return m.iat[0] if not m.empty else ""

    g = df.assign(_doc=df["documento"].astype(str).str.strip().replace("", "(sem CPF/CNPJ)"))
    out = (g.groupby("_doc").agg(Credor=("credor", _maiscomum), Qtd=("id", "count"),
                                 Total=("valor_num", "sum"))
           .reset_index().rename(columns={"_doc": "CPF/CNPJ"})
           .sort_values("Total", ascending=False).head(n).reset_index(drop=True))
    return out[cols]


def aging_vencidos(df: pd.DataFrame, hoje=None) -> pd.DataFrame:
    """Faixas de atraso (apenas vencidos: vencimento < hoje)."""
    cols = ["Faixa", "Qtd", "Total"]
    if df.empty or "vencimento_dt" not in df.columns:
        return pd.DataFrame(columns=cols)
    hoje = pd.Timestamp(hoje or pd.Timestamp.now().date())
    venc = df["vencimento_dt"]
    atraso = (hoje - venc).dt.days
    venc_mask = venc.notna() & (atraso > 0)
    d = df[venc_mask].assign(_atraso=atraso[venc_mask])
    if d.empty:
        return pd.DataFrame(columns=cols)
    faixas = [(1, 7, "1–7 dias"), (8, 15, "8–15 dias"), (16, 30, "16–30 dias"),
              (31, 60, "31–60 dias"), (61, 90, "61–90 dias"), (91, 10**9, "90+ dias")]
    linhas = []
    for ini, fim, nome in faixas:
        sub = d[(d["_atraso"] >= ini) & (d["_atraso"] <= fim)]
        if len(sub):
            linhas.append({"Faixa": nome, "Qtd": int(len(sub)),
                           "Total": float(sub["valor_num"].sum())})
    return pd.DataFrame(linhas, columns=cols)


def kpis(df: pd.DataFrame, hoje=None) -> dict:
    if df.empty:
        return {"total": 0.0, "qtd": 0, "ticket": 0.0, "venc_qtd": 0, "venc_total": 0.0}
    hoje = pd.Timestamp(hoje or pd.Timestamp.now().date())
    total = float(df["valor_num"].sum())
    qtd = int(len(df))
    venc_total = 0.0
    venc_qtd = 0
    if "vencimento_dt" in df.columns:
        venc = df["vencimento_dt"]
        m = venc.notna() & (venc < hoje)
        venc_qtd = int(m.sum())
        venc_total = float(df.loc[m, "valor_num"].sum())
    return {"total": total, "qtd": qtd, "ticket": total / qtd if qtd else 0.0,
            "venc_qtd": venc_qtd, "venc_total": venc_total}


# ---------------------------------------------------------------------------
# EXPORTAÇÃO XLSX
# ---------------------------------------------------------------------------
def _auto_largura(ws, max_linhas: int = 200, larg_max: int = 60):
    from openpyxl.utils import get_column_letter
    fim = min(ws.max_row, max_linhas)
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=fim), start=1):
        maxlen = 0
        for c in col_cells:
            if c.value is not None:
                maxlen = max(maxlen, len(str(c.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, maxlen + 2), larg_max)


def gerar_xlsx(titulo: str, resumo: dict, abas: dict, dados: pd.DataFrame) -> bytes:
    """resumo: dict de KPIs; abas: {nome: DataFrame}; dados: linhas detalhadas."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        cab = pd.DataFrame([{"Relatório": titulo}])
        cab.to_excel(xw, sheet_name="Resumo", index=False, startrow=0)
        res_df = pd.DataFrame(list(resumo.items()), columns=["Indicador", "Valor"])
        res_df.to_excel(xw, sheet_name="Resumo", index=False, startrow=2)
        for nome, dfx in abas.items():
            (dfx if dfx is not None else pd.DataFrame()).to_excel(
                xw, sheet_name=nome[:31], index=False)
        if dados is not None and not dados.empty:
            dados.to_excel(xw, sheet_name="Dados", index=False)
        for ws in xw.sheets.values():
            _auto_largura(ws)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# EXPORTAÇÃO PDF (tabular, via reportlab)
# ---------------------------------------------------------------------------
def gerar_pdf(titulo: str, subtitulo: str, kpi_linhas: list, tabelas: dict,
              analitico: dict | None = None) -> bytes:
    """
    kpi_linhas: lista de (rótulo, valor_str).
    tabelas: {nome: (colunas:list[str], linhas:list[list[str]])}.
    analitico (opcional): {"titulo": str, "colunas": list[str],
                           "grupos": [(nome_grupo, linhas, rodape_str), ...],
                           "rodape": str} — impresso após quebra de página, com
                           uma tabela por grupo (linhas quebram entre páginas).
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, PageBreak)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.2 * cm, bottomMargin=1.2 * cm,
                            leftMargin=1.2 * cm, rightMargin=1.2 * cm, title=titulo)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=16, spaceAfter=4)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, spaceBefore=10, spaceAfter=4)
    h3 = ParagraphStyle("h3", parent=styles["Heading3"], fontSize=10, spaceBefore=8, spaceAfter=2)
    el = [Paragraph(titulo, h1), Paragraph(subtitulo, sub), Spacer(1, 0.3 * cm)]

    AZUL = colors.HexColor("#0a7d2c")
    CINZA = colors.HexColor("#f0f2f6")

    if kpi_linhas:
        kt = Table([[Paragraph(f"<b>{k}</b>", styles["Normal"]), v] for k, v in kpi_linhas],
                   colWidths=[6 * cm, 11 * cm])
        kt.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, CINZA]),
            ("FONTSIZE", (0, 0), (-1, -1), 9), ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
        el += [kt, Spacer(1, 0.2 * cm)]

    for nome, (colunas, linhas) in tabelas.items():
        el.append(Paragraph(nome, h2))
        if not linhas:
            el.append(Paragraph("Sem dados.", sub)); continue
        data = [colunas] + linhas
        ncol = len(colunas)
        if ncol == 3:                       # dimensão, Qtd, Total
            larguras = [11 * cm, 2.4 * cm, 4 * cm]
        elif ncol == 4:                     # CPF/CNPJ, Credor, Qtd, Total
            larguras = [3.4 * cm, 8 * cm, 2 * cm, 4 * cm]
        else:
            larguras = [17.4 / ncol * cm] * ncol
        t = Table(data, colWidths=larguras, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), AZUL),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CINZA]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ("ALIGN", (-2, 1), (-2, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2)]))
        el += [t, Spacer(1, 0.2 * cm)]

    # ===== Seção analítica (linha a linha, agrupada) =====
    if analitico and analitico.get("grupos"):
        el.append(PageBreak())
        el.append(Paragraph(analitico.get("titulo", "Relatório Analítico das Despesas"), h2))
        colunas = analitico["colunas"]
        # ID · Venc. · Credor · Descrição · Tipo · Valor
        larguras = [1.7 * cm, 1.6 * cm, 4.3 * cm, 5.6 * cm, 2.3 * cm, 1.9 * cm]
        if len(colunas) != len(larguras):
            larguras = [17.4 / len(colunas) * cm] * len(colunas)
        # células de texto com QUEBRA DE LINHA (string pura não quebra e um texto
        # comprido "vaza" por cima da coluna vizinha)
        cel = ParagraphStyle("cel", parent=styles["Normal"], fontSize=6, leading=7)
        _QUEBRA = {2, 3, 4}                  # Credor, Descrição, Tipo
        for nome_grupo, linhas, rodape_grp in analitico["grupos"]:
            el.append(Paragraph(nome_grupo, h3))
            data = [colunas]
            for ln in linhas:
                data.append([Paragraph(str(v), cel) if i in _QUEBRA else str(v)
                             for i, v in enumerate(ln)])
            t = Table(data, colWidths=larguras, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), AZUL),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, 0), 7),
                ("FONTSIZE", (0, 1), (-1, -1), 6),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CINZA]),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 1.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5)]))
            el.append(t)
            if rodape_grp:
                el.append(Paragraph(f"<b>{rodape_grp}</b>", sub))
            el.append(Spacer(1, 0.15 * cm))
        if analitico.get("rodape"):
            el.append(Spacer(1, 0.2 * cm))
            el.append(Paragraph(f"<b>{analitico['rodape']}</b>", styles["Normal"]))

    doc.build(el)
    return buf.getvalue()


def analitico_despesas(df: pd.DataFrame, max_linhas: int = 1500,
                       agrupar_por: str = "centro_custo",
                       titulo: str | None = None):
    """Despesas linha a linha agrupadas pela dimensão `agrupar_por` (padrão Centro
    de Custo; ex.: 'grupo_lote'), com subtotal por grupo e total geral, prontas
    para o PDF. Retorna dict no formato do parâmetro 'analitico' de gerar_pdf, ou
    None se df vazio. Aplica teto de linhas (avisa no rodapé quando cortar)."""
    if df is None or df.empty or agrupar_por not in df.columns:
        return None
    d = df.copy()
    d["_cc"] = d[agrupar_por].astype(str).str.strip().replace("", VAZIO)
    d["_v"] = pd.to_numeric(d["valor_num"], errors="coerce").fillna(0.0)
    if "vencimento_dt" in d.columns:
        d = d.sort_values(["_cc", "vencimento_dt", "id"])
    else:
        d = d.sort_values(["_cc", "vencimento", "id"])

    def _trunc(s, n):
        s = str(s or "").strip()
        return s if len(s) <= n else s[: n - 1] + "…"

    grupos, usadas, cortadas = [], 0, 0
    total = float(d["_v"].sum())
    qtd_total = len(d)
    for cc, g in d.groupby("_cc", sort=True):
        sub = float(g["_v"].sum())
        if usadas >= max_linhas:
            cortadas += len(g)
            grupos.append((f"{cc} — {len(g)} lançamento(s) · Subtotal {_moeda(sub)}",
                           [["…", "", "(grupo omitido por limite de páginas)", "", "", ""]],
                           ""))
            continue
        linhas = []
        for _, r in g.iterrows():
            if usadas >= max_linhas:
                cortadas += 1
                continue
            linhas.append([str(r.get("id", "")), _trunc(r.get("vencimento", ""), 10),
                           _trunc(r.get("credor", ""), 60), _trunc(r.get("descricao", ""), 95),
                           _trunc(r.get("tipo_despesa", ""), 26), _moeda(r.get("_v", 0))])
            usadas += 1
        grupos.append((f"{cc} — {len(g)} lançamento(s) · Subtotal {_moeda(sub)}",
                       linhas, ""))
    rodape = f"Total geral: {_moeda(total)} · {qtd_total} lançamento(s)"
    if cortadas:
        rodape += (f" · ⚠️ {cortadas} linha(s) omitida(s) no PDF (limite de {max_linhas}; "
                   f"use o XLSX para o detalhamento completo)")
    return {"titulo": titulo or "Relatório Analítico das Despesas (por Centro de Custo)",
            "colunas": ["ID", "Venc.", "Credor", "Descrição", "Tipo", "Valor"],
            "grupos": grupos, "rodape": rodape}


def fluxo_despesas(df: pd.DataFrame) -> dict | None:
    """Fluxo de despesas por VENCIMENTO com granularidade automática conforme o
    intervalo dos dados: até ~35 dias -> Diário; até ~180 dias -> Semanal;
    acima -> Mensal. Retorna {'nivel', 'tabela'(Período|Qtd|Total)} ou None."""
    if df is None or df.empty or "vencimento_dt" not in df.columns:
        return None
    d = df.copy()
    d["_v"] = pd.to_numeric(d["valor_num"], errors="coerce").fillna(0.0)
    d = d.dropna(subset=["vencimento_dt"])
    if d.empty:
        return None
    span = (d["vencimento_dt"].max() - d["vencimento_dt"].min()).days
    if span <= 35:
        nivel = "Diário"
        d["_per"] = d["vencimento_dt"].dt.strftime("%d/%m/%Y")
        d["_ord"] = d["vencimento_dt"]
    elif span <= 180:
        nivel = "Semanal"
        ini = d["vencimento_dt"] - pd.to_timedelta(d["vencimento_dt"].dt.weekday, unit="D")
        d["_per"] = "Semana de " + ini.dt.strftime("%d/%m/%Y")
        d["_ord"] = ini
    else:
        nivel = "Mensal"
        d["_per"] = d["vencimento_dt"].dt.strftime("%m/%Y")
        d["_ord"] = d["vencimento_dt"].dt.to_period("M").dt.start_time
    tab = (d.groupby(["_ord", "_per"], as_index=False)
             .agg(Qtd=("_v", "size"), Total=("_v", "sum"))
             .sort_values("_ord"))
    tab = tab.rename(columns={"_per": "Período"})[["Período", "Qtd", "Total"]]
    return {"nivel": nivel, "tabela": tab}


# ---------------------------------------------------------------------------
# Tabelas dinâmicas (reproduzem painéis da planilha)
# ---------------------------------------------------------------------------
_FORMA_ORDER = ["BeeVale", "Boleto", "Pix", "Transferência Bancária", "Somapay"]


def pagamentos_diarios(df: pd.DataFrame) -> dict:
    """
    Quantidade de pagamentos por dia (data de pagamento, col X) e forma (col J).
    Considera só lançamentos com data de pagamento preenchida (= pagos).
    Retorna {tabela, formas, contagem, media}.
    """
    vazio = {"tabela": pd.DataFrame(columns=["Data Pgt", "Qtd"]),
             "formas": [], "contagem": {}, "media": {}}
    if df.empty or "data_pagamento_dt" not in df.columns:
        return vazio
    d = df[df["data_pagamento_dt"].notna()].copy()
    if d.empty:
        return vazio
    d["forma_pagamento"] = d["forma_pagamento"].astype(str).str.strip().replace("", "—")
    piv = d.pivot_table(index="data_pagamento_dt", columns="forma_pagamento",
                        values="id", aggfunc="count", fill_value=0)
    formas = ([f for f in _FORMA_ORDER if f in piv.columns]
              + [f for f in piv.columns if f not in _FORMA_ORDER])
    piv = piv[formas].astype(int)
    piv["Qtd"] = piv.sum(axis=1)
    cols_tot = formas + ["Qtd"]
    contagem = piv[cols_tot].sum(axis=0).astype(int)
    dias_ativos = (piv[cols_tot] > 0).sum(axis=0).replace(0, 1)
    media = (contagem / dias_ativos).round(0).astype(int)
    piv = piv.sort_index(ascending=False)
    tab = piv.reset_index().rename(columns={"data_pagamento_dt": "Data Pgt"})
    tab["Data Pgt"] = pd.to_datetime(tab["Data Pgt"]).dt.strftime("%d/%m/%Y")
    return {"tabela": tab, "formas": formas,
            "contagem": contagem.to_dict(), "media": media.to_dict()}


def necessidade_caixa(df: pd.DataFrame, hoje: pd.Timestamp,
                      dias_antes: int = 15, dias_depois: int = 20) -> dict:
    """
    Necessidade de caixa por dia (vencimento) e conta de pagamento (col U),
    janela D-`dias_antes` a D+`dias_depois`. Só lançamentos a pagar
    (não Pago e não Cancelado). Retorna {tabela, contas, vencido, rodape}.
    """
    vazio = {"tabela": pd.DataFrame(columns=["Vencimento", "Soma"]),
             "contas": [], "vencido": [], "rodape": pd.DataFrame()}
    if df.empty or "vencimento_dt" not in df.columns:
        return vazio
    status = df["status_pgt"].astype(str).str.strip().str.lower()
    d = df[(~status.isin(["pago", "cancelado"])) & df["vencimento_dt"].notna()].copy()
    ini = hoje - pd.Timedelta(days=dias_antes)
    fim = hoje + pd.Timedelta(days=dias_depois)
    d = d[(d["vencimento_dt"] >= ini) & (d["vencimento_dt"] <= fim)]
    if d.empty:
        return vazio
    d["conta_fmt"] = d["conta_fmt"].astype(str).str.strip().replace("", "DEFINIR")
    piv = d.pivot_table(index="vencimento_dt", columns="conta_fmt",
                        values="valor_num", aggfunc="sum", fill_value=0.0)
    contas = list(piv.columns)
    piv = piv.sort_index(ascending=True)
    piv["Soma"] = piv[contas].sum(axis=1)
    cols_tot = contas + ["Soma"]
    vencido = [bool(x) for x in (piv.index < hoje)]
    total = piv[cols_tot].sum(axis=0)
    atrasado = piv.loc[piv.index < hoje, cols_tot].sum(axis=0)
    vence_hoje = piv.loc[piv.index == hoje, cols_tot].sum(axis=0)
    rodape = pd.DataFrame([total, atrasado, vence_hoje],
                          index=["Somatório Total", "Atrasado (vencido)", "Vence hoje"])
    tab = piv.reset_index().rename(columns={"vencimento_dt": "Vencimento"})
    tab["Vencimento"] = pd.to_datetime(tab["Vencimento"]).dt.strftime("%d/%m/%Y")
    return {"tabela": tab, "contas": contas, "vencido": vencido, "rodape": rodape}