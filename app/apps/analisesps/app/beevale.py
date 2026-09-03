# -*- coding: utf-8 -*-
"""
beevale.py — geração de planilhas BeeVale.

Parte 1 (esta): "Planilha Cadastro BeeVale" — recebe e-mails/CPFs colados,
extrai os CPFs, busca os dados na planilha 'Dados Documentos' e gera o XLSX de
cadastro para download. Portado de `gerarCadastroBeeValePorLinhasErro`.
"""
from __future__ import annotations

import io
import re

# Fonte dos dados de cadastro (planilha 'Dados Documentos').
DADOS_DOCS_ID = "1fqi4QUOVGUd1_4Gg4vK5qP_IMOSgFaw8DD9MDgmM3vo"
DADOS_DOCS_ABA = "Dados Documentos"

# Pasta do Drive onde os xlsx por card são salvos (mesma do Apps Script).
DRIVE_FOLDER_ID = "1ycGeXKyHhU5R4Y242eT13MkGQ85CI6gv"


def only_digits(s) -> str:
    return re.sub(r"\D+", "", str(s or ""))


def extrair_cpfs(texto: str) -> list:
    """Extrai CPFs de um texto: primeiro de e-mails 11díg@bwsconstrucoes.com.br,
    senão de números soltos de 11 dígitos. Mantém ordem e remove duplicados."""
    s = str(texto or "")
    achados = re.findall(r"(\d{11})@bwsconstrucoes\.com\.br", s, flags=re.I)
    if not achados:
        achados = re.findall(r"\b\d{11}\b", s)
    vistos, out = set(), []
    for c in achados:
        if c not in vistos:
            vistos.add(c)
            out.append(c)
    return out


def normaliza_cpf11(v) -> str:
    if v in (None, ""):
        return ""
    s = str(v).strip()
    if re.fullmatch(r"\d+(\.0+)?", s):
        s = re.sub(r"\.0+$", "", s)
    s = only_digits(s)
    if not s:
        return ""
    return s.rjust(11, "0")[-11:]


def formata_cpf(cpf) -> str:
    d = only_digits(cpf)
    if len(d) != 11:
        return str(cpf or "")
    return f"{d[0:3]}.{d[3:6]}.{d[6:9]}-{d[9:11]}"


def nome_impresso(nome: str) -> str:
    p = [x for x in str(nome or "").strip().split() if x]
    if not p:
        return ""
    if len(p) == 1:
        return p[0]
    return f"{p[0]} {p[-1]}"


def data_br(v) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", s):
        return s
    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    m = re.fullmatch(r"(\d{2})-(\d{2})-(\d{4})", s)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
    return s


def telefone_br(phone) -> str:
    d = only_digits(phone)
    if not d:
        return ""
    if len(d) >= 12 and d.startswith("55"):
        d = d[2:]
    if len(d) > 11:
        d = d[-11:]
    if len(d) == 11:
        return f"({d[0:2]}) {d[2:7]}-{d[7:]}"
    if len(d) == 10:
        return f"({d[0:2]}) {d[2:6]}-{d[6:]}"
    return str(phone or "")


def registro_de_linha(nome, nascimento, telefone, cpf11) -> dict:
    cpfd = only_digits(cpf11)
    return {
        "nomeCompleto": str(nome or "").strip(),
        "nomeImpresso": nome_impresso(nome),
        "cpf": formata_cpf(cpf11),
        "email": cpfd + "@bwsconstrucoes.com.br",
        "dataNascimento": data_br(nascimento),
        "celular": telefone_br(telefone),
    }


def buscar_registros_por_cpf(cpfs: list):
    """Lê 'Dados Documentos' e devolve (encontrados, nao_encontrados).
    Precisa de internet (service account com acesso à planilha)."""
    import gsheets
    ws = gsheets._abrir_aba(DADOS_DOCS_ABA, planilha_id=DADOS_DOCS_ID)
    vals = ws.get_all_values()
    if not vals:
        return [], list(cpfs)
    h = vals[0]
    i_nome = gsheets._idx_hdr(h, "Nome Completo")
    i_nasc = gsheets._idx_hdr(h, "Data de Nascimento")
    i_tel = gsheets._idx_hdr(h, "Telefone Celular", "Telefone", "Celular")
    i_cpfn = gsheets._idx_hdr(h, "CPF Números")
    i_cpff = gsheets._idx_hdr(h, "CPF (Cadastro de Pessoa Física)", "CPF")

    faltando = [k for k, v in {"Nome Completo": i_nome, "Data de Nascimento": i_nasc,
                               "Telefone": i_tel, "CPF": (i_cpfn if i_cpfn is not None
                                                          else i_cpff)}.items() if v is None]
    if faltando:
        raise RuntimeError("Cabeçalhos não encontrados em 'Dados Documentos': "
                           + ", ".join(faltando))

    desejados = set(cpfs)
    achados = {}

    def _g(r, i):
        return r[i].strip() if (i is not None and i < len(r)) else ""

    for r in vals[1:]:
        cpf = normaliza_cpf11(_g(r, i_cpfn)) or normaliza_cpf11(_g(r, i_cpff))
        if not cpf or cpf not in desejados:
            continue
        achados[cpf] = registro_de_linha(_g(r, i_nome), _g(r, i_nasc), _g(r, i_tel), cpf)

    encontrados = [achados[c] for c in cpfs if c in achados]
    nao = [c for c in cpfs if c not in achados]
    return encontrados, nao


def cadastro_xlsx(registros: list) -> bytes:
    """Gera o XLSX de Cadastro BeeVale (mesmo layout do script)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Cadastro BeeVale"
    headers = ["Nome completo", "Nome Impresso no Cartão", "CPF", "Email",
               "Data de nascimento", "Celular"]
    ws.append(headers)
    for r in registros:
        ws.append([r.get("nomeCompleto", ""), r.get("nomeImpresso", ""), r.get("cpf", ""),
                   r.get("email", ""), r.get("dataNascimento", ""), r.get("celular", "")])
    for col in range(1, 7):
        letra = get_column_letter(col)
        for cell in ws[letra]:
            cell.number_format = "@"
        ws.column_dimensions[letra].width = 24
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_valor_br(s) -> float:
    s = str(s or "").strip()
    if not s:
        return 0.0
    try:
        return float(s.replace(".", "").replace(",", "."))
    except ValueError:
        return 0.0


def pagamento_xlsx(linhas: list) -> bytes:
    """XLSX de Pagamento BeeVale. linhas: [{nome,email,valor_num,cpf,cardId}]."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamento BeeVale"
    headers = ["Nome", "Email", "Carteira", "Benefício", "Valor", "Tipo de Recarga",
               "Dias úteis", "Documento de Identificação", "Nome Completo (Colaborador)",
               "Centro de Custo", "Categoria"]
    ws.append(headers)
    for l in linhas:
        ws.append([l.get("nome", ""), l.get("email", ""), "Produção", "Livre",
                   float(l.get("valor_num", 0) or 0), "Mensal", 0, l.get("cpf", ""),
                   l.get("nome", ""), str(l.get("cardId", "")), "Terceirizados"])
    for c in (1, 2, 3, 4, 6, 8, 9, 10, 11):       # texto
        for cell in ws[get_column_letter(c)]:
            cell.number_format = "@"
    for cell in ws[get_column_letter(5)]:         # Valor
        cell.number_format = "0.00"
    for cell in ws[get_column_letter(7)]:         # Dias úteis
        cell.number_format = "0"
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def montar_descricao(descricao_atual, pagamento_link: str, cadastro_link: str) -> str:
    """Replica buildUpdatedDescription_ do Apps Script: remove linhas antigas de
    'Pagamento BeeVale:', 'Cadastro BeeVale:' e 'Planilha BeeVale:' e põe as novas
    no TOPO, preservando o restante da descrição."""
    pag = f"Pagamento BeeVale: {pagamento_link}"
    cad = f"Cadastro BeeVale: {cadastro_link}"
    texto = str(descricao_atual or "").strip()
    mantidas = []
    for l in texto.split("\n"):
        ll = l.strip()
        if (re.match(r"(?i)^Pagamento BeeVale:\s*", ll)
                or re.match(r"(?i)^Cadastro BeeVale:\s*", ll)
                or re.match(r"(?i)^Planilha BeeVale:\s*", ll)):
            continue
        mantidas.append(l)
    texto = "\n".join(mantidas).strip()
    topo = f"{pag}\n{cad}"
    return f"{topo}\n\n{texto}" if texto else topo


def gerar_beevale(ids: list) -> dict:
    """
    Para cada card BeeVale (igual ao Apps Script):
      1) busca CPF/valor no Pipefy + dados do colaborador na database;
      2) gera Pagamento e Cadastro (xlsx, UM par por card);
      3) sobe os dois no Drive (pasta DRIVE_FOLDER_ID), link público;
      4) atualiza a DESCRIÇÃO do card (links no topo) + Documentação Fiscal='BeeVale'.
    Retorna {ok_ids, erros, resultados, atualizados}. Precisa de internet/token e
    de acesso de escrita da Service Account à pasta do Drive (idealmente Shared Drive).
    """
    import pipefy
    import drive
    from datetime import datetime

    ids = [str(i) for i in ids]
    cards = pipefy.fetch_cards(ids)

    erros, card_cpf, cpfs = [], {}, []
    for cid in ids:
        c = cards.get(cid)
        if not c:
            erros.append({"Card": cid, "Motivo": "Card não retornado pela API."})
            continue
        cpf = pipefy.extract_cpf(c["fieldsById"].get(pipefy.FIELD_CADASTRO, ""))
        if not cpf:
            erros.append({"Card": cid, "Motivo": "Campo 'Cadastro BeeVale' vazio/inválido."})
            continue
        card_cpf[cid] = cpf
        cpfs.append(cpf)

    cpfs_unicos = list(dict.fromkeys(cpfs))
    records = pipefy.fetch_beevale_records(cpfs_unicos) if cpfs_unicos else {}
    hoje = datetime.now().strftime("%d.%m.%Y")

    ok_ids, updates, resultados = [], [], []
    for cid in ids:
        cpf = card_cpf.get(cid)
        if not cpf:
            continue
        rec = records.get(cpf)
        if not rec:
            erros.append({"Card": cid, "Motivo": f"CPF {cpf} não encontrado na database BeeVale."})
            continue
        nome = (rec.get("nome_completo") or "").strip()
        if not nome:
            erros.append({"Card": cid, "Motivo": f"Registro sem Nome Completo (CPF {cpf})."})
            continue
        cpf_mask = (rec.get("cpf") or cpf).strip()
        valor = parse_valor_br(cards[cid]["fieldsById"].get(pipefy.FIELD_VALOR, ""))
        email = only_digits(cpf_mask) + "@bwsconstrucoes.com.br"
        try:
            pag_bytes = pagamento_xlsx([{
                "nome": nome, "email": email, "valor_num": valor,
                "cpf": formata_cpf(cpf_mask), "cardId": cid}])
            cad_bytes = cadastro_xlsx([registro_de_linha(
                nome, rec.get("data_de_nascimento", ""),
                rec.get("telefone_celular", ""), cpf_mask)])
            up_pag = drive.upload_xlsx_publico(
                pag_bytes, f"{cid}_{hoje}.pagamento.beevale.xlsx", DRIVE_FOLDER_ID)
            up_cad = drive.upload_xlsx_publico(
                cad_bytes, f"{cid}_{hoje}.cadastro.beevale.xlsx", DRIVE_FOLDER_ID)
        except Exception as e:
            erros.append({"Card": cid, "Motivo": f"Falha ao gerar/subir no Drive: {e}"})
            continue

        desc_atual = cards[cid]["fieldsById"].get(pipefy.FIELD_DESCRICAO, "")
        nova_desc = montar_descricao(desc_atual, up_pag["link"], up_cad["link"])
        updates.append({"cardId": cid, "descricao": nova_desc})
        resultados.append({"Card": cid, "Pagamento": up_pag["link"],
                           "Cadastro": up_cad["link"]})
        ok_ids.append(cid)

    atualizados = []
    if updates:
        try:
            falhas = pipefy.atualizar_descricao_e_doc_fiscal(updates)
        except Exception as e:
            falhas = [u["cardId"] for u in updates]
            erros.append({"Card": "(lote)", "Motivo": f"Arquivos no Drive OK, mas falhou "
                                                       f"o update dos cards: {e}"})
        atualizados = [u["cardId"] for u in updates if u["cardId"] not in falhas]
        for cid in falhas:
            erros.append({"Card": cid, "Motivo": "Arquivos no Drive OK, mas falhou "
                                                 "atualizar descrição/doc fiscal do card."})

    return {"ok_ids": ok_ids, "erros": erros, "resultados": resultados,
            "atualizados": atualizados}
