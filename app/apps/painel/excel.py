# -*- coding: utf-8 -*-
"""
O relatório em Excel — um arquivo, várias abas.

A primeira versão desta conversão exportava CSV para não acrescentar
biblioteca. Era pior do que parecia: o relatório original tinha **oito abas num
arquivo só**, e em CSV isso vira oito downloads soltos que ninguém junta depois.
O dono pediu Excel; `openpyxl` custa 250 KB e resolve.

Não usa `pandas`: escrever célula a célula com o `openpyxl` gasta muito menos
memória, e memória é o recurso escasso deste serviço.

Cada aba sai formatada — cabeçalho fixo, coluna de dinheiro em R$, largura
proporcional ao conteúdo, filtro no topo. É o que separa "abriu" de "dá para
trabalhar nisso".
"""
from __future__ import annotations

import datetime as dt
import io

FORMATO_MOEDA = 'R$ #,##0.00;[Red]-R$ #,##0.00'
FORMATO_DATA = "DD/MM/YYYY"
FORMATO_PCT = '0.0"%"'

# Colunas cujo nome indica dinheiro — recebem o formato de moeda sem precisar
# de configuração por aba.
DINHEIRO = ("valor", "pago", "receb", "aberto", "pagar", "bruto", "líquid",
            "liquid", "retid", "total", "executado", "comprometido", "juros",
            "multa", "desconto", "saldo", "quota", "base", "rateio",
            "resultado", "despesa", "receita", "crédito", "credito", "encargo")


def _e_dinheiro(titulo: str) -> bool:
    t = (titulo or "").lower()
    return any(p in t for p in DINHEIRO) and "%" not in t


def _largura(titulo: str, valores) -> int:
    maior = len(str(titulo))
    for v in valores[:200]:          # amostra: 200 linhas bastam para calibrar
        maior = max(maior, len(str(v if v is not None else "")))
    return min(max(maior + 3, 11), 52)


def montar(abas, titulo_arquivo: str = "Relatório") -> bytes:
    """`abas` é uma lista de (nome, colunas, linhas):
    `colunas` são pares (chave, título) e `linhas` são dicionários.

    Aba vazia entra assim mesmo, com um aviso dentro — some da lista seria pior:
    quem abre fica sem saber se não havia dado ou se o relatório falhou.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    livro = Workbook()
    livro.remove(livro.active)

    cabecalho_fundo = PatternFill("solid", fgColor="12385C")
    cabecalho_fonte = Font(color="FFFFFF", bold=True, size=10)

    for nome, colunas, linhas in abas:
        # o Excel recusa nome de aba com mais de 31 caracteres ou com : \ / ? * [ ]
        seguro = nome[:31]
        for proibido in ':\\/?*[]':
            seguro = seguro.replace(proibido, "-")
        folha = livro.create_sheet(seguro)

        if not linhas:
            folha["A1"] = f"(sem dados para {nome} nos filtros escolhidos)"
            folha["A1"].font = Font(italic=True, color="5A6883")
            folha.column_dimensions["A"].width = 60
            continue

        for coluna, (_chave, rotulo) in enumerate(colunas, start=1):
            celula = folha.cell(row=1, column=coluna, value=rotulo)
            celula.fill = cabecalho_fundo
            celula.font = cabecalho_fonte
            celula.alignment = Alignment(vertical="center", wrap_text=False)

        for numero, linha in enumerate(linhas, start=2):
            for coluna, (chave, rotulo) in enumerate(colunas, start=1):
                valor = linha.get(chave)
                if isinstance(valor, dt.datetime):
                    valor = valor.date()
                celula = folha.cell(row=numero, column=coluna, value=valor)
                if isinstance(valor, dt.date):
                    celula.number_format = FORMATO_DATA
                elif isinstance(valor, (int, float)) and not isinstance(valor, bool):
                    celula.number_format = (FORMATO_PCT if "%" in rotulo
                                            else FORMATO_MOEDA if _e_dinheiro(rotulo)
                                            else "#,##0")

        for coluna, (chave, rotulo) in enumerate(colunas, start=1):
            folha.column_dimensions[get_column_letter(coluna)].width = _largura(
                rotulo, [l.get(chave) for l in linhas])

        # cabeçalho fixo e filtro: numa aba de 20 mil linhas isso é o que
        # permite trabalhar em vez de só olhar
        folha.freeze_panes = "A2"
        folha.auto_filter.ref = (
            f"A1:{get_column_letter(len(colunas))}{len(linhas) + 1}")

    buffer = io.BytesIO()
    livro.save(buffer)
    return buffer.getvalue()


def nome_do_arquivo(assunto: str) -> str:
    return f"{assunto}_{dt.date.today():%Y-%m-%d}.xlsx"


# ---------------------------------------------------------------------------
# As colunas de cada assunto
# ---------------------------------------------------------------------------
# Os títulos são os que aparecem na tela: quem baixa o arquivo tem de reconhecer
# as colunas sem precisar traduzir nada.
COLUNAS = {
    "dre": [("linha", "Linha"), ("executado", "Executado"),
            ("aberto", "Em aberto"), ("comprometido", "Comprometido")],
    "despesas": [("nome", "Grupo ou categoria"), ("valor", "Valor"),
                 ("pct_total", "% do total")],
    "credores": [("nome", "Credor"), ("pago", "Já pago"),
                 ("aberto", "A pagar"), ("titulos", "Títulos")],
    "analitico": [("data", "Data (pagto ou vencto)"),
                  ("data_vencimento", "Vencimento"),
                  ("data_pagamento", "Pagamento"),
                  ("atraso", "Atraso (dias)"),
                  ("credor", "Credor"), ("cnpj", "CNPJ/CPF"),
                  ("grupo", "Grupo"), ("categoria", "Categoria"),
                  ("obra", "Obra"), ("projeto", "Projeto"),
                  ("documento", "Documento"), ("observacao", "Observação"),
                  ("conta", "Conta corrente"), ("situacao", "Situação"),
                  ("vencimento", "Situação do vencimento"),
                  ("pedido", "Pedido de compra"), ("medicao", "Medição"),
                  ("lancamento", "Nº no OMIE"),
                  ("pago", "Pago"), ("a_pagar", "A pagar"),
                  ("juros", "Juros"), ("multa", "Multa"), ("total", "Total")],
    "medicoes": [("medicao", "Medição"), ("cliente", "Cliente"),
                 ("obra", "Obra"), ("projeto", "Projeto"),
                 ("documento", "Documento"), ("data", "Data"),
                 ("bruto", "Bruto"), ("recebido", "Recebido"),
                 ("retido", "Retido na fonte"), ("a_receber", "A receber"),
                 ("situacao", "Situação")],
    "outras": [("categoria", "Categoria"), ("recebido", "Recebido"),
               ("a_receber", "A receber"), ("titulos", "Títulos")],
    "fluxo": [("rotulo", "Mês"), ("entradas", "Entradas"), ("saidas", "Saídas"),
              ("liquido", "No mês"), ("acumulado", "Acumulado")],
    "obras": [("nome", "Projeto ou obra"), ("receita", "Receita líquida"),
              ("despesa", "Despesas"), ("resultado", "Resultado")],
    "execucao": [("nome", "Projeto ou obra"), ("executado", "Executado"),
                 ("a_executar", "A executar"), ("comprometido", "Comprometido"),
                 ("pct", "% andado")],
    "quotas": [("socio", "Sócio"), ("tipo", "Tipo"), ("projeto", "Projeto"),
               ("pct", "%"), ("base", "Base de cálculo"),
               ("credito_bws", "Crédito BWS"), ("quota", "Quota"),
               ("visao", "Como foi calculado")],
    "posicao": [("socio", "Sócio"), ("tipo", "Tipo"), ("projetos", "Projetos"),
                ("quota", "Quota"), ("ajustes", "Ajustes"), ("saldo", "Saldo")],
    # Aportes: cada recorte tem a sua aba, como na tela.
    "aporte_socio": [("socio", "Sócio ou parceiro"), ("aportado", "Aportado"),
                     ("devolvido", "Devolvido"), ("saldo", "Saldo"),
                     ("pct", "% do aportado"), ("lancamentos", "Lançamentos")],
    "aporte_obra": [("obra", "Obra"), ("socio", "Sócio ou parceiro"),
                    ("aportado", "Aportado"), ("devolvido", "Devolvido"),
                    ("saldo", "Saldo"), ("falta", "Falta p/ igualar"),
                    ("lancamentos", "Lançamentos")],
    "aporte_tipo": [("obra", "Obra"), ("tipo", "Tipo"), ("aportado", "Aportado"),
                    ("devolvido", "Devolvido"), ("saldo", "Saldo"),
                    ("lancamentos", "Lançamentos")],
    "aporte_lancamentos": [("data", "Data"), ("obra", "Obra"),
                           ("socio", "Sócio ou parceiro"), ("tipo", "Tipo"),
                           ("categoria", "Categoria"), ("valor", "Valor"),
                           ("conta", "Conta corrente"),
                           ("documento", "Documento"),
                           ("observacao", "Observação")],
    "aporte_dividendos": [("socio", "Sócio"), ("pago", "Pago a ele"),
                          ("recebido", "Recebido dele"), ("liquido", "Líquido"),
                          ("lancamentos", "Lançamentos")],
    "divisao": [("obra", "Obra"), ("resultado", "Resultado realizado"),
                ("dividendos", "Dividendos pagos"), ("disponivel", "Disponível")],
}
