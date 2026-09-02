# -*- coding: utf-8 -*-
"""
De-para Departamento (obra no OMIE) -> Projeto.

Le a planilha Google "Bases de Dados Pipefy", aba "C. Diarios":
    coluna AJ = codigo do departamento OMIE (cCodDep)
    coluna AK = Projeto
e grava na tabela painel.depto_projeto.

Diferencas em relacao a versao que rodava no PC:
  - a credencial vem de GOOGLE_CREDENTIALS_BASE64, o padrao do repositorio,
    em vez de um arquivo credenciais.json ao lado do script;
  - o identificador da planilha vem de PAINEL_SHEET_PROJETOS, em vez de estar
    escrito no codigo;
  - le apenas a faixa AJ:AK, nunca a aba inteira (regra de memoria do CONTEXTO
    secao 3.7: nada de get_all_values em aba grande).
"""
from __future__ import annotations

import os
import csv
import logging
import datetime as dt

log = logging.getLogger("painel.projetos")

ABA = "C. Diários"
FAIXA = f"'{ABA}'!AJ:AK"          # AJ = cCodDep, AK = Projeto
ESCOPOS = ["https://www.googleapis.com/auth/spreadsheets.readonly"]

# indices 0-based das colunas AJ e AK num arquivo exportado inteiro
COL_AJ = 35  # A=0 ... AJ=35
COL_AK = 36  # AK=36

# Conferencia conhecida da especificacao: este departamento e do projeto CEIFOR7.
DEPTO_CONFERENCIA, PROJETO_CONFERENCIA = "583764198", "CEIFOR7"


def _s(v):
    return "" if v is None else str(v).strip()


def id_da_planilha() -> str:
    sid = os.getenv("PAINEL_SHEET_PROJETOS", "").strip()
    if not sid:
        raise RuntimeError(
            "PAINEL_SHEET_PROJETOS nao configurada. E o identificador da planilha "
            "'Bases de Dados Pipefy' (o trecho da URL entre /d/ e /edit).")
    return sid


# --------------------------------------------------------------------------- 
# Leitura via Google Sheets API (service account)
# --------------------------------------------------------------------------- 
def ler_via_api(sheet_id: str | None = None, faixa: str = FAIXA):
    import json
    from base64 import b64decode
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build

    b64 = os.getenv("GOOGLE_CREDENTIALS_BASE64", "")
    if not b64:
        raise RuntimeError("GOOGLE_CREDENTIALS_BASE64 nao configurado.")
    info = json.loads(b64decode(b64).decode("utf-8"))
    creds = Credentials.from_service_account_info(info, scopes=ESCOPOS)
    servico = build("sheets", "v4", credentials=creds, cache_discovery=False)
    resp = servico.spreadsheets().values().get(
        spreadsheetId=sheet_id or id_da_planilha(), range=faixa,
        valueRenderOption="UNFORMATTED_VALUE").execute()
    return resp.get("values", [])  # lista de linhas [AJ, AK]


# --------------------------------------------------------------------------- 
# Leitura offline de arquivo exportado (.xlsx / .csv) — so para conferencia local
# --------------------------------------------------------------------------- 
def ler_de_arquivo(caminho):
    """Le AJ e AK de um arquivo exportado. Aceita a aba inteira (recorta pela
    posicao das colunas) ou um arquivo com so duas colunas."""
    ext = os.path.splitext(caminho)[1].lower()
    linhas = []
    if ext == ".csv":
        with open(caminho, newline="", encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                linhas.append(row)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb[ABA] if ABA in wb.sheetnames else wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            linhas.append(list(row))
        wb.close()

    largura = max((len(r) for r in linhas[:50]), default=0)
    out = []
    for r in linhas:
        if largura > 2:
            aj = r[COL_AJ] if len(r) > COL_AJ else None
            ak = r[COL_AK] if len(r) > COL_AK else None
        else:
            aj = r[0] if len(r) > 0 else None
            ak = r[1] if len(r) > 1 else None
        out.append([aj, ak])
    return out


# --------------------------------------------------------------------------- 
# Monta o mapa cCodDep -> projeto
# --------------------------------------------------------------------------- 
def construir_mapa(linhas):
    """Ignora cabecalho e vazios. Em conflito (mesmo depto com projetos
    diferentes), o primeiro vence e o caso vai para o log."""
    mapa, conflitos = {}, []
    for i, linha in enumerate(linhas):
        aj = linha[0] if len(linha) > 0 else None
        ak = linha[1] if len(linha) > 1 else None
        dep, proj = _s(aj), _s(ak)
        if not dep or not proj:
            continue
        if i == 0 and not dep.replace(".", "").isdigit():
            continue
        if dep in mapa:
            if mapa[dep] != proj:
                conflitos.append((dep, mapa[dep], proj))
            continue
        mapa[dep] = proj
    return mapa, conflitos


def gravar(conn, mapa):
    agora = dt.datetime.now().isoformat(timespec="seconds")
    conn.executemany(
        "INSERT INTO depto_projeto (ccoddep, projeto, sync_em) VALUES (?,?,?) "
        "ON CONFLICT(ccoddep) DO UPDATE SET projeto=excluded.projeto, sync_em=excluded.sync_em",
        [(d, p, agora) for d, p in mapa.items()])
    conn.commit()
    return len(mapa)


def sincronizar(conn, de_arquivo=None):
    """Le (API ou arquivo), monta o mapa e grava. Devolve a quantidade de obras."""
    linhas = ler_de_arquivo(de_arquivo) if de_arquivo else ler_via_api()
    mapa, conflitos = construir_mapa(linhas)
    for dep, a, b in conflitos[:10]:
        log.warning("Obra %s com projetos divergentes na planilha: %s vs %s (mantido %s).",
                    dep, a, b, a)
    n = gravar(conn, mapa)
    if DEPTO_CONFERENCIA in mapa and mapa[DEPTO_CONFERENCIA] != PROJETO_CONFERENCIA:
        log.warning("Conferencia: obra %s esperava projeto %s, veio '%s'.",
                    DEPTO_CONFERENCIA, PROJETO_CONFERENCIA, mapa[DEPTO_CONFERENCIA])
    return n
